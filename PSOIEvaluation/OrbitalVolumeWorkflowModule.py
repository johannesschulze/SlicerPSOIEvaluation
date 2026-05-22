"""
OrbitalVolumeWorkflowModule
============================
Slicer-Modul mit grafischer Oberfläche für:
  1. Erstellung der Orbita-Eingangsebene aus einer Closed Curve
     (Logik aus orbital_surface_from_curve.py)
  2. Segmentierung des intraorbitalen Volumens via Fast Marching
     (Logik aus intraorbital_volume_segmentation.py)

Unterstützt linke und rechte Orbita getrennt.
"""

import logging
import os
import time
from typing import Optional

import numpy as np
import vtk
import ctk

from __main__ import slicer
import qt

from slicer.i18n import tr as _
from slicer.i18n import translate
from slicer.ScriptedLoadableModule import (
    ScriptedLoadableModule,
    ScriptedLoadableModuleLogic,
    ScriptedLoadableModuleTest,
    ScriptedLoadableModuleWidget,
)
from slicer.parameterNodeWrapper import parameterNodeWrapper
from slicer.util import VTKObservationMixin
from slicer import (
    vtkMRMLScalarVolumeNode, vtkMRMLModelNode, vtkMRMLSegmentationNode,
    vtkMRMLMarkupsFiducialNode, vtkMRMLLinearTransformNode,
)

import SimpleITK as sitk
import sitkUtils

# ═══════════════════════════════════════════════════════════════════════════════
# Modul-Klasse
# ═══════════════════════════════════════════════════════════════════════════════

class OrbitalVolumeWorkflowModule(ScriptedLoadableModule):

    def __init__(self, parent):
        ScriptedLoadableModule.__init__(self, parent)
        self.parent.title = _("OrbitalVolumeWorkflowModule")
        self.parent.categories = [translate("qSlicerAbstractCoreModule", "PSOI Evaluation")]
        self.parent.dependencies = []
        self.parent.contributors = ["Johannes Schulze (Bundeswehrkrankenhaus Ulm)"]
        _doc_html = os.path.join(
            os.path.dirname(__file__), "Resources", "Docs", "OrbitalVolumeWorkflowModule","orbita_volume_workflow_DE.html"
        )
        _doc_url = "file://" + _doc_html.replace("\\", "/")
        self.parent.helpText = (
            _("This module creates an entry plane for the orbit from a closed curve "
              "and subsequently segments the intraorbital volume using Fast Marching. "
              "The left and right orbits are managed separately.")
            + f' <a target="_blank" href="{_doc_url}">'
            + _("Open manual (DE)")
            + "</a>"
        )
        self.parent.acknowledgementText = _(
            "Developed by Johannes Schulze (Bundeswehrkrankenhaus Ulm/Universität Ulm) "
            "without external funding."
        )


# ═══════════════════════════════════════════════════════════════════════════════
# Parameter-Node
# ═══════════════════════════════════════════════════════════════════════════════

@parameterNodeWrapper
class OrbitalVolumeWorkflowModuleParameterNode:
    ctVolume:              vtkMRMLScalarVolumeNode
    planeModelLeft:        vtkMRMLModelNode
    planeModelRight:       vtkMRMLModelNode
    segmentationNodeLeft:  vtkMRMLSegmentationNode
    segmentationNodeRight: vtkMRMLSegmentationNode
    seedNodeLeft:          vtkMRMLMarkupsFiducialNode
    seedNodeRight:         vtkMRMLMarkupsFiducialNode
    step:             int  = 0
    sideIsLeft:       bool = True
    rightSideVisited: bool = False
    # Segmentierungsparameter – separat für linke und rechte Seite
    huMinLeft:          float   = -200
    huMinRight:         float   = -200
    huMaxLeft:          float   =  300
    huMaxRight:         float   =  300
    autoSeedLeft:       bool  = True
    autoSeedRight:      bool  = True
    seedOffsetLeft:     float = 10.0
    seedOffsetRight:    float = 10.0
    radiusMarginLeft:   float =  2.0
    radiusMarginRight:  float =  2.0
    stoppingValueLeft:  float = 25.0
    stoppingValueRight: float = 25.0
    speedSigmaLeft:     float = 70.0
    speedSigmaRight:    float = 70.0
    posteriorBoostLeft:  float = 2.5
    posteriorBoostRight: float = 2.5
    showSeedLeft:           bool  = False
    showSeedRight:          bool  = False
    # 0 = manual, 1 = mirror contralateral, 2 = model-based
    seedModeLeft:            int   = 0
    seedModeRight:           int   = 0
    # 'fastmarching' | 'threshold' | ''
    segMethodLeft:           str   = ""
    segMethodRight:          str   = ""
    modelSeedNodeLeft:       vtkMRMLSegmentationNode
    modelSeedNodeRight:      vtkMRMLSegmentationNode
    modelSeedTransformLeft:  vtkMRMLLinearTransformNode
    modelSeedTransformRight: vtkMRMLLinearTransformNode
    posteriorCutoffNode:     vtkMRMLMarkupsFiducialNode   # shared, beide Seiten
    rimLandmarkNodeLeft:     vtkMRMLMarkupsFiducialNode
    rimLandmarkNodeRight:    vtkMRMLMarkupsFiducialNode
    removeSatellitesLeft:   bool  = True
    removeSatellitesRight:  bool  = True
    satelliteDiamLeft:      float = 3.0
    satelliteDiamRight:     float = 3.0
    contraPositionedNodeLeft:       vtkMRMLSegmentationNode
    contraPositionedNodeRight:      vtkMRMLSegmentationNode
    contraPositionedTransformLeft:  vtkMRMLLinearTransformNode
    contraPositionedTransformRight: vtkMRMLLinearTransformNode
    currentStep : str = ""


# ═══════════════════════════════════════════════════════════════════════════════
# Presets
# ═══════════════════════════════════════════════════════════════════════════════

# Keys match combobox item text; "Manual" has no entry (never applied programmatically).
ORBITAL_RIM_LANDMARKS = [
    ("L1",  "Supraorbitalforamen / Incisura supraorbitalis",
             "Übergang mediales Drittel / laterale zwei Drittel des Supraorbitalrands"),
    ("L2",  "Mitte Supraorbitalrand",
             "Scheitelpunkt des Supraorbitalrands zwischen L1 und L3"),
    ("L3",  "Frontozygomatiksutur",
             "Lateraler Orbitarand, Höhe Lateralkanthus"),
    ("L4",  "Mitte lateraler Orbitarand",
             "Maximale Lateralprojektion des Os zygomaticum"),
    ("L5",  "Mitte inferiolateraler Orbitarand",
             "Mitte des Segments L4–L6, Übergang lateral → inferior"),
    ("L6",  "Sutura zygomaticomaxillaris",
             "Übergang lateraler → inferiorer Rand"),
    ("L7",  "Mitte inferiorer Orbitarand",
             "Mitte des Segments L6–L8"),
    ("L8",  "Mitte inferomedial / Lacrimalkante inferior",
             "Mitte des Segments L7–L9, Übergang inferior → medial"),
    ("L9",  "Frontomaxillarsutur / Lacrimalkante superior",
             "Übergang medialer → superomedial verlaufender Rand"),
    ("L10", "Mitte superomedial",
             "Mitte des Segments L9–L1, Region der Trochlea"),
]

SEGMENTATION_PRESETS = {
    "CT Bone Window": {
        "huMin": -300, "huMax": 600,
        "stoppingValue": 35.0, "speedSigma": 100.0,
    },
    "Intraoperative CBCT": {
        "huMin": 0, "huMax": 600,
        "stoppingValue": 20.0, "speedSigma": 150.0,
    },
}


# ═══════════════════════════════════════════════════════════════════════════════
# Widget
# ═══════════════════════════════════════════════════════════════════════════════

class OrbitalVolumeWorkflowModuleWidget(ScriptedLoadableModuleWidget, VTKObservationMixin):

    def __init__(self, parent=None) -> None:
        ScriptedLoadableModuleWidget.__init__(self, parent)
        VTKObservationMixin.__init__(self)
        self.logic = None
        self._parameterNode = None
        self._parameterNodeGuiTag = None
        # Per-ParameterNode-Zustand (gespeichert nach Node-ID, damit Wechsel verlustfrei sind)
        self._statePerPN = {}   # {pn_id: {curveNodes, surfaceTexts, segTexts, segNodes, segIds}}
        # Aktiv genutzter Per-Seite-Zustand (True = links, False = rechts)
        self._curveNodes   = {True: None, False: None}
        self._surfaceTexts = {True: "",   False: ""}
        self._segTexts     = {True: "",   False: ""}
        self._segVolumes   = {True: None, False: None}  # volume_ml per side
        self._vrDisplayNode = None
        # Segmentierungs-Nodes und Observer für automatische Volumen-Neuberechnung
        self._segNodes     = {True: None, False: None}
        self._segIds       = {True: None, False: None}
        self._segObservers         = {True: None, False: None}
        self._cutoffMarkupObserver = None
        self._landmarkObservers    = {True: None, False: None}  # (node, [obs_tags…]) per side
        self._volumeUpdateSide = None
        self._applyingPreset = False
        self._volumeUpdateTimer = qt.QTimer()
        self._volumeUpdateTimer.setSingleShot(True)
        self._volumeUpdateTimer.setInterval(1500)
        self._volumeUpdateTimer.connect('timeout()', self._doVolumeUpdate)

    # ------------------------------------------------------------------
    # Lebenszyklus
    # ------------------------------------------------------------------

    def setup(self) -> None:
        # Load translation for current locale
        _translator = qt.QTranslator()
        _locale = qt.QLocale.system().name()  # e.g. "de_DE"
        _qm = self.resourcePath(f"Translations/OrbitalVolumeWorkflowModule_{_locale}.qm")
        if _translator.load(_qm):
            qt.QCoreApplication.installTranslator(_translator)
        else:
            # Try language-only fallback (e.g. "de" from "de_DE")
            _qm_lang = self.resourcePath(
                f"Translations/OrbitalVolumeWorkflowModule_{_locale.split('_')[0]}.qm"
            )
            if _translator.load(_qm_lang):
                qt.QCoreApplication.installTranslator(_translator)

        ScriptedLoadableModuleWidget.setup(self)

        # QTextBrowser intercepts all link clicks; redirect them to the system browser.
        for browser in self.parent.findChildren(qt.QTextBrowser):
            browser.setOpenLinks(False)
            browser.anchorClicked.connect(lambda url: qt.QDesktopServices.openUrl(url))

        uiWidget = slicer.util.loadUI(self.resourcePath("UI/OrbitalVolumeWorkflowModule.ui"))
        self.layout.addWidget(uiWidget)
        self.ui = slicer.util.childWidgetVariables(uiWidget)
        uiWidget.setMRMLScene(slicer.mrmlScene)
        self._uiWidget = uiWidget

        slicer.app.connect("paletteChanged(QPalette)", self._refreshButtonStyles)
        slicer.app.connect("styleChanged(QString)", self._refreshButtonStyles)


        self.logic = OrbitalVolumeWorkflowModuleLogic()

        self.addObserver(slicer.mrmlScene, slicer.mrmlScene.StartCloseEvent, self.onSceneStartClose)
        self.addObserver(slicer.mrmlScene, slicer.mrmlScene.EndCloseEvent,   self.onSceneEndClose)

        # ParameterNode-Selektor: erst Filter setzen, dann Scene zuweisen, damit
        # der Filter beim initialen Scene-Scan bereits aktiv ist.
        self.ui.parameterNodeSelector.addAttribute(
            "vtkMRMLScriptedModuleNode", "ModuleName", self.moduleName
        )
        self.ui.parameterNodeSelector.setMRMLScene(slicer.mrmlScene)
        self.ui.parameterNodeSelector.connect(
            "currentNodeChanged(vtkMRMLNode*)", self.setParameterNode
        )

        # Workflow-Buttons
        self.ui.createSurfaceButton.connect("clicked(bool)",    self.onCreateSurfaceButton)
        self.ui.fastMarchingButton.connect("clicked(bool)", self.onFastMarchingButton)
        self.ui.thresholdButton.connect("clicked(bool)",    self.onThresholdButton)
        self.ui.clearSideButton.connect("clicked(bool)",       self.onClearSideButton)
        self.ui.exportResultsButton.connect("clicked(bool)",   self.onExportResultsButton)
        self.ui.autoSeedCheckBox.connect("toggled(bool)",       self.onAutoSeedToggled)
        #self.ui.stepsToolbox.connect("currentChanged(int)",     self.onStepsToolboxCurrentChanged)
       
        self.ui.placeCutoffButton.connect("clicked(bool)", self.onPlaceCutoffButton)
        self.ui.placeCutoffButton.setIcon(qt.QIcon(self.resourcePath("Icons/MarkupsFiducialMouseModePlace.png")))

        self.ui.selectIslandButton.connect("clicked(bool)", self.onSelectIslandButton)
        self.ui.removeExtrusionsButton.connect("clicked(bool)", self.onRemoveExtrusionsButton)
        self.ui.performCutoffButton.connect("clicked(bool)", self.onPerformCutoffButton)
        self.ui.refreshVolumeButton.connect("clicked(bool)", self.onRefreshVolumeButton)

        # Open Documentation Button
        self.ui.openDocumentationButton.connect("clicked(bool)", self.onOpenDocumentationClicked)

        self.ui.cutoffNodeSelector.connect(
            "currentNodeChanged(vtkMRMLNode*)", self.onCutoffNodeSelectorChanged
        )
        self.ui.removeSatellitesCheckBox.connect(
            "toggled(bool)", lambda checked: self.ui.satelliteDiamSpinBox.setEnabled(checked)
        )
        for rb in [self.ui.rbSeedManual, self.ui.rbSeedContralateral, self.ui.rbSeedModelBased]:
            rb.connect("toggled(bool)", lambda checked: checked and self.onSeedModeChanged())
        self.ui.modelSeedSelector.connect(
            "currentNodeChanged(vtkMRMLNode*)", self.onModelSeedChanged
        )
        self.ui.loadTemplateButton.connect("clicked(bool)",  self.onLoadTemplateButton)
        self.ui.positionModelButton.connect("clicked(bool)",  self.onPositionModelButton)
        self.ui.mirrorTemplateButton.connect("toggled(bool)", self.onMirrorTemplateToggled)
        self.ui.positionContralateralButton.connect("clicked(bool)", self.onPositionContralateralButton)

        # Seiten-Buttons
        self.ui.btnSideLeft.connect("clicked()",  lambda: self.onSideChanged(True))
        #self.ui.btnSideLeft.setIcon(qt.QIcon(self.resourcePath("Icons/orbit_left.png")))
        self.ui.btnSideRight.connect("clicked()", lambda: self.onSideChanged(False))
        #self.ui.btnSideRight.setIcon(qt.QIcon(self.resourcePath("Icons/orbit_right.png")))

        # Collapsibles
        self.ui.pageLandmarks.connect("clicked(bool)", lambda x: self.onPageCollapsibleClicked(self.ui.pageLandmarks, x))
        self.ui.pageOrbitalSurface.connect("clicked(bool)", lambda x: self.onPageCollapsibleClicked(self.ui.pageOrbitalSurface, x))
        self.ui.pageVolumeSegmentation.connect("clicked(bool)", lambda x: self.onPageCollapsibleClicked(self.ui.pageVolumeSegmentation, x))

        # Selektoren manuell beobachten (kein SlicerParameterName für seitenspezifische Nodes)
        self.ui.curveSelector.connect(
            "currentNodeChanged(vtkMRMLNode*)", self.onCurveChanged)
        self.ui.planeModelSelector.connect(
            "currentNodeChanged(vtkMRMLNode*)", self.onPlaneModelChanged)

        # Preset combobox
        self.ui.presetComboBox.connect("currentIndexChanged(int)", self.onPresetChanged)
        self.ui.huMinSpinBox.connect("valueChanged(int)", self._onPresetSpinboxChanged)
        self.ui.huMaxSpinBox.connect("valueChanged(int)", self._onPresetSpinboxChanged)
        self.ui.stoppingValueSpinBox.connect("valueChanged(double)", self._onPresetSpinboxChanged)
        self.ui.speedSigmaSpinBox.connect("valueChanged(double)", self._onPresetSpinboxChanged)

        # CT-Volume → Volume-Rendering; HU-Shift-Slider; Toggle-Button
        self.ui.ctVolumeSelector.connect(
            "currentNodeChanged(vtkMRMLNode*)", self.onCTVolumeChanged)
        self.ui.huShiftSlider.connect("valueChanged(double)", self.onHUShiftChanged)
        self.ui.toggleVolumeRenderingButton.connect(
            "toggled(bool)", self.onToggleVolumeRendering)
        self.ui.resampleCTButton.connect("clicked(bool)", self.onResampleCTButton)

        # Landmark workflow buttons
        self.ui.newLandmarkNodeButton.connect("clicked(bool)", self.onNewLandmarkNodeButton)
        self.ui.landmarkStartButton.connect("clicked(bool)",   self.onStartLandmarkButton)
        self.ui.landmarkPrevButton.connect("clicked(bool)",    self.onLandmarkPrevButton)
        self.ui.curveFromLandmarksButton.connect("clicked(bool)", self.onCurveFromLandmarksButton)
        self.ui.landmarkNodeSelector.connect(
            "currentNodeChanged(vtkMRMLNode*)", self.onLandmarkNodeSelectorChanged
        )

        # Neue-Kurve-Button
        self.ui.createCurveButton.connect("clicked(bool)", self.onCreateCurveButton)
        self.ui.createCurveButton.setIcon(qt.QIcon(self.resourcePath("Icons/MarkupsClosedCurveMouseModePlace.png")))

        # Manuelle Segmentierungsknoten-Auswahl
        self.ui.segNodeSelector.setMRMLScene(slicer.mrmlScene)
        self.ui.segNodeSelector.connect(
            "currentNodeChanged(vtkMRMLNode*)", self.onSegNodeSelectorChanged)

        # Segment-Editor-Button (nach Segmentierung aktiviert)
        self.ui.openSegmentEditorButton.connect("clicked(bool)", self.onOpenSegmentEditorButton)

        self.initializeParameterNode()

    def cleanup(self) -> None:
        self._volumeUpdateTimer.stop()
        for side in [True, False]:
            node = self._segNodes.get(side)
            obs  = self._segObservers.get(side)
            if node is not None and obs is not None:
                node.RemoveObserver(obs)
            lm_entry = self._landmarkObservers.get(side)
            if lm_entry is not None:
                lm_node, lm_obs_list = lm_entry
                for obs in lm_obs_list:
                    try:
                        lm_node.RemoveObserver(obs)
                    except Exception:
                        pass
        if self._parameterNode is not None:
            cutoff_node = getattr(self._parameterNode, "posteriorCutoffNode", None)
            if cutoff_node is not None and self._cutoffMarkupObserver is not None:
                cutoff_node.RemoveObserver(self._cutoffMarkupObserver)
        self.removeObservers()

    def enter(self) -> None:
        self.initializeParameterNode()
        if self._parameterNode:
            self._activateCurrentStep()
            isLeft = self._parameterNode.sideIsLeft
            self.ui.btnSideLeft.setChecked(isLeft)
            self.ui.btnSideRight.setChecked(not isLeft)
            self._refreshSideUI(isLeft)

    def exit(self) -> None:
        if self._parameterNode:
            self._parameterNode.disconnectGui(self._parameterNodeGuiTag)
            self._parameterNodeGuiTag = None

    def onSceneStartClose(self, caller, event) -> None:
        self.setParameterNode(None)

    def onSceneEndClose(self, caller, event) -> None:
        if self.parent.isEntered:
            self.initializeParameterNode()

    def _findWidgetByName(self, name):
        for i in range(self.ui.mainVerticalLayout.count()):
            item_widget = self.ui.mainVerticalLayout.itemAt(i).widget()

            if item_widget != None and item_widget.objectName == name:
                return item_widget
        
        return None

    # ------------------------------------------------------------------
    # Preset
    # ------------------------------------------------------------------

    def onPresetChanged(self, index: int) -> None:
        if self._applyingPreset:
            return
        name = self.ui.presetComboBox.itemText(index)
        preset = SEGMENTATION_PRESETS.get(name)
        if preset is None:
            return  # "Manual" selected programmatically or by user — nothing to apply
        ok = slicer.util.confirmOkCancelDisplay(
            _("Apply preset \"{name}\"?\n\n"
              "HU min = {huMin}, HU max = {huMax}\n"
              "Stopping value = {stoppingValue} mm, Speed sigma = {speedSigma} HU\n\n"
              "This will overwrite the current values.").format(name=name, **preset),
            _("Apply Preset"),
        )
        if not ok:
            # Revert combobox without triggering this handler again
            self._applyingPreset = True
            self.ui.presetComboBox.setCurrentIndex(
                self.ui.presetComboBox.count - 1  # "Manual"
            )
            self._applyingPreset = False
            return
        self._applyingPreset = True
        self.ui.huMinSpinBox.setValue(preset["huMin"])
        self.ui.huMaxSpinBox.setValue(preset["huMax"])
        self.ui.stoppingValueSpinBox.setValue(preset["stoppingValue"])
        self.ui.speedSigmaSpinBox.setValue(preset["speedSigma"])
        self._applyingPreset = False

    def _onPresetSpinboxChanged(self, _value) -> None:
        if self._applyingPreset:
            return
        # Switch to "Manual" without triggering onPresetChanged
        self._applyingPreset = True
        manual_index = self.ui.presetComboBox.count - 1
        self.ui.presetComboBox.setCurrentIndex(manual_index)
        self._applyingPreset = False

    def onOpenDocumentationClicked(self, _value) -> None:
        import webbrowser
        from os import path

        _locale = qt.QLocale().system().name()[3:]
        filename = self.resourcePath(f"Docs/OrbitalVolumeWorkflowModule/orbita_volume_workflow_{_locale}.html")

        if not os.path.exists(filename):
            filename = self.resourcePath(f"Docs/OrbitalVolumeWorkflowModule/orbita_volume_workflow.html")
        
        webbrowser.open("file://" + filename)

    def onPageCollapsibleClicked(self, sender: qt.QWidget, expanded: bool) -> None:
        collapsibleWidgets = self._uiWidget.findChildren(ctk.ctkCollapsibleButton)

        for w in collapsibleWidgets:
            w.setChecked(False)

        if expanded:
            sender.setChecked(True)
            self._parameterNode.currentStep = sender.objectName
        else:
            self._parameterNode.currentStep = ""

        self._refreshButtonStyles()
            

    def _activateCurrentStep(self):
        currentStep = self._parameterNode.currentStep
        page_widget = self._findWidgetByName(currentStep)

        if page_widget != None:
            self.onPageCollapsibleClicked(page_widget, True)

        
    # ------------------------------------------------------------------
    # Parameter-Node
    # ------------------------------------------------------------------

    def initializeParameterNode(self) -> None:
        if self._parameterNode is not None:
            # Re-entering after exit(): GUI was disconnected but observers and
            # self._parameterNode are still intact — just reconnect the GUI.
            wasBlocked = self.ui.ctVolumeSelector.blockSignals(True)
            self.ui.ctVolumeSelector.setCurrentNode(self._parameterNode.ctVolume)
            self.ui.ctVolumeSelector.blockSignals(wasBlocked)
            self._parameterNodeGuiTag = self._parameterNode.connectGui(self.ui)
            wasBlocked = self.ui.parameterNodeSelector.blockSignals(True)
            self.ui.parameterNodeSelector.setCurrentNode(self._parameterNode.parameterNode)
            self.ui.parameterNodeSelector.blockSignals(wasBlocked)
            return
        # Nach einem Modul-Reload: zuletzt aktiven PN aus den App-Settings wiederherstellen.
        last_id = slicer.app.userSettings().value("OrbitalVolumeWorkflowModule/LastParameterNodeID", "")
        restored = None
        if last_id:
            restored = slicer.mrmlScene.GetNodeByID(last_id)
        if restored is not None:
            pn = OrbitalVolumeWorkflowModuleParameterNode(restored)
        else:
            pn = self.logic.getParameterNode()
        # Direkt setzen – nicht auf das Combo-Box-Signal warten (Box kann beim
        # ersten Start leer/inaktiv sein, bevor ein Node existiert).
        self.setParameterNode(pn)
        # Combo Box synchronisieren ohne erneutes Signal
        wasBlocked = self.ui.parameterNodeSelector.blockSignals(True)
        self.ui.parameterNodeSelector.setCurrentNode(pn.parameterNode)
        self.ui.parameterNodeSelector.blockSignals(wasBlocked)

    def setParameterNode(self, inputParameterNode) -> None:
        # Eingabe kann ein roher vtkMRMLNode (aus Combo Box) oder bereits
        # ein gewrappter OrbitalVolumeWorkflowModuleParameterNode sein.
        if isinstance(inputParameterNode, slicer.vtkMRMLNode):
            # ModuleName-Attribut setzen, damit der Node im gefilterten Selector bleibt
            inputParameterNode.SetAttribute("ModuleName", self.moduleName)
            inputParameterNode = OrbitalVolumeWorkflowModuleParameterNode(inputParameterNode)

        # Gleicher unterliegender Node → nichts zu tun
        if (self._parameterNode is not None
                and inputParameterNode is not None
                and self._parameterNode.parameterNode is inputParameterNode.parameterNode):
            return

        # ---- Alten PN aufräumen ----
        if self._parameterNode is not None:
            # Save current GUI values so they survive a PN switch
            self._saveParamsForSide(self._parameterNode.sideIsLeft)
            old_id = self._parameterNode.parameterNode.GetID()
            # Per-PN-Zustand sichern
            self._statePerPN[old_id] = {
                "curveNodes":   dict(self._curveNodes),
                "surfaceTexts": dict(self._surfaceTexts),
                "segTexts":     dict(self._segTexts),
                "segNodes":     dict(self._segNodes),
                "segIds":       dict(self._segIds),
                "segVolumes":   dict(self._segVolumes),
            }
            # Observer vom alten Seg-Nodes lösen
            for side in [True, False]:
                node = self._segNodes.get(side)
                obs  = self._segObservers.get(side)
                if node is not None and obs is not None:
                    node.RemoveObserver(obs)
            self._segObservers = {True: None, False: None}
            cutoff_node = getattr(self._parameterNode, "posteriorCutoffNode", None)
            if cutoff_node is not None and self._cutoffMarkupObserver is not None:
                cutoff_node.RemoveObserver(self._cutoffMarkupObserver)
            self._cutoffMarkupObserver = None
            # Landmark-Observer vom alten PN lösen
            for side in [True, False]:
                lm_entry = self._landmarkObservers.get(side)
                if lm_entry is not None:
                    lm_node, lm_obs_list = lm_entry
                    for obs in lm_obs_list:
                        try:
                            lm_node.RemoveObserver(obs)
                        except Exception:
                            pass
            self._landmarkObservers = {True: None, False: None}
            self._parameterNode.disconnectGui(self._parameterNodeGuiTag)
            self._parameterNodeGuiTag = None

        self._parameterNode = inputParameterNode

        if self._parameterNode is None:
            return

        # ---- Neuen PN aktivieren ----
        new_id = self._parameterNode.parameterNode.GetID()
        # ID für Reload-Wiederherstellung in den App-Settings speichern
        slicer.app.userSettings().setValue("OrbitalVolumeWorkflowModule/LastParameterNodeID", new_id)
        if new_id in self._statePerPN:
            s = self._statePerPN[new_id]
            self._curveNodes   = s["curveNodes"]
            self._surfaceTexts = s["surfaceTexts"]
            self._segTexts     = s["segTexts"]
            self._segNodes     = s["segNodes"]
            self._segIds       = s["segIds"]
            self._segVolumes   = s.get("segVolumes", {True: None, False: None})
        else:
            self._curveNodes   = {True: None, False: None}
            self._surfaceTexts = {True: "",   False: ""}
            self._segTexts     = {True: "",   False: ""}
            self._segNodes     = {True: None, False: None}
            self._segIds       = {True: None, False: None}
            self._segVolumes   = {True: None, False: None}

        # Observer an vorhandene Seg-Nodes des neuen PN hängen
        self._segObservers = {True: None, False: None}
        for isLeft, attr in [(True, "segmentationNodeLeft"), (False, "segmentationNodeRight")]:
            seg_node = getattr(self._parameterNode, attr, None)
            if seg_node is None:
                continue
            # Segment-ID aus dem Node holen (wird u.U. neu befüllt)
            seg_id = self._segIds.get(isLeft)
            if seg_id is None:
                seg = seg_node.GetSegmentation()
                for i in range(seg.GetNumberOfSegments()):
                    if seg.GetNthSegment(i).GetName() == "IntraorbitalVolume":
                        seg_id = seg.GetNthSegmentID(i)
                        self._segIds[isLeft] = seg_id
                        break
            if seg_id:
                self._segNodes[isLeft] = seg_node
                self._segObservers[isLeft] = seg_node.AddObserver(
                    vtk.vtkCommand.ModifiedEvent,
                    lambda c, e, side=isLeft: self._onSegmentModified(side),
                )

        # Observer an vorhandenen Cutoff-Markup-Node hängen + Selector synchronisieren
        cutoff_node = getattr(self._parameterNode, "posteriorCutoffNode", None)
        if cutoff_node is not None:
            self._cutoffMarkupObserver = cutoff_node.AddObserver(
                vtk.vtkCommand.ModifiedEvent,
                lambda c, e: self._onCutoffMarkupModified(),
            )
        wasBlocked = self.ui.cutoffNodeSelector.blockSignals(True)
        self.ui.cutoffNodeSelector.setCurrentNode(cutoff_node)
        self.ui.cutoffNodeSelector.blockSignals(wasBlocked)
        self._onCutoffMarkupModified()

        # Landmark-Observer an vorhandene Nodes des neuen PN hängen
        self._landmarkObservers = {True: None, False: None}
        for isLeft, attr in [(True, "rimLandmarkNodeLeft"), (False, "rimLandmarkNodeRight")]:
            lm_node = getattr(self._parameterNode, attr, None)
            if lm_node is not None:
                obs_defined = lm_node.AddObserver(
                    slicer.vtkMRMLMarkupsNode.PointPositionDefinedEvent,
                    lambda c, e, side=isLeft: self._onLandmarkPointAdded(side),
                )
                obs_removed = lm_node.AddObserver(
                    slicer.vtkMRMLMarkupsNode.PointRemovedEvent,
                    lambda c, e, side=isLeft: self._updateLandmarkUI(side),
                )
                self._landmarkObservers[isLeft] = (lm_node, [obs_defined, obs_removed])

        # ctVolumeSelector: sync to new PN's volume; no signal blocking so that
        # onCTVolumeChanged fires and the display updates correctly.
        self.ui.ctVolumeSelector.setCurrentNode(self._parameterNode.ctVolume)

        self._parameterNodeGuiTag = self._parameterNode.connectGui(self.ui)

        isLeft = self._parameterNode.sideIsLeft
        self.ui.btnSideLeft.setChecked(isLeft)
        self.ui.btnSideRight.setChecked(not isLeft)
        self._refreshSideUI(isLeft)

    # ------------------------------------------------------------------
    # Signalhandler – Seitenauswahl
    # ------------------------------------------------------------------

    def onSideChanged(self, isLeft: bool) -> None:
        if self._parameterNode:
            # Aktuelle Parameter der verlassenen Seite speichern
            self._saveParamsForSide(self._parameterNode.sideIsLeft)
            # Beim ersten Besuch der rechten Seite: linke Parameter übernehmen
            if not isLeft and not self._parameterNode.rightSideVisited:
                self._copyParamsToSide(fromLeft=True)
                self._parameterNode.rightSideVisited = True
            self._parameterNode.sideIsLeft = isLeft
        self.ui.btnSideLeft.setChecked(isLeft)
        self.ui.btnSideRight.setChecked(not isLeft)
        self._refreshSideUI(isLeft)

    def _refreshSideUI(self, isLeft: bool) -> None:
        """Aktualisiert Selektoren, Parameter-Widgets und Labels für die gewählte Seite."""
        self.ui.curveSelector.blockSignals(True)
        self.ui.curveSelector.setCurrentNode(self._curveNodes[isLeft])
        self.ui.curveSelector.blockSignals(False)

        # refresh button palettes after side switch
        btnActive = self.ui.btnSideLeft if isLeft else self.ui.btnSideRight
        btnInactive = self.ui.btnSideRight if isLeft else self.ui.btnSideLeft

        self._refreshButtonStyles()

        self.ui.planeModelSelector.blockSignals(True)
        if self._parameterNode:
            plane = (self._parameterNode.planeModelLeft
                     if isLeft else self._parameterNode.planeModelRight)
            self.ui.planeModelSelector.setCurrentNode(plane)
        self.ui.planeModelSelector.blockSignals(False)

        self.ui.segNodeSelector.blockSignals(True)
        if self._parameterNode:
            seg = (self._parameterNode.segmentationNodeLeft if isLeft
                   else self._parameterNode.segmentationNodeRight)
            self.ui.segNodeSelector.setCurrentNode(seg)
        self.ui.segNodeSelector.blockSignals(False)

        self._loadParamsForSide(isLeft)
        self.ui.createSurfaceButton.setEnabled(self._curveNodes[isLeft] is not None)
        self.ui.surfaceResultLabel.setText(self._surfaceTexts[isLeft])

        # Landmark selector
        if self._parameterNode:
            lm_node = (self._parameterNode.rimLandmarkNodeLeft if isLeft
                       else self._parameterNode.rimLandmarkNodeRight)
            self.ui.landmarkNodeSelector.blockSignals(True)
            self.ui.landmarkNodeSelector.setCurrentNode(lm_node)
            self.ui.landmarkNodeSelector.blockSignals(False)
        self._updateLandmarkUI(isLeft)

        # Volumen nach Reload aus vorhandenem Segmentierungsknoten nachmessen
        if not self._segTexts[isLeft] and self._parameterNode is not None:
            seg_node = (self._parameterNode.segmentationNodeLeft if isLeft
                        else self._parameterNode.segmentationNodeRight)
            if seg_node is not None:
                seg = seg_node.GetSegmentation()
                seg_id = None
                for i in range(seg.GetNumberOfSegments()):
                    if seg.GetNthSegment(i).GetName() == "IntraorbitalVolume":
                        seg_id = seg.GetNthSegmentID(i)
                        break
                if seg_id is not None:
                    try:
                        vol_ml, voxel_count = self._calculateVolumeFromSegment(seg_node, seg_id)
                        self._segVolumes[isLeft] = vol_ml
                        self._segTexts[isLeft] = (
                            _("<b>Intraorbital volume: {vol:.2f} ml</b>"
                              " &nbsp;<i>(reloaded)</i><br>"
                              "Voxels: {vox}").format(
                                vol=vol_ml, vox=f"{voxel_count:,}")
                        )
                    except Exception:
                        pass

        self.ui.segmentationResultLabel.setText(self._segTexts[isLeft])

    def _saveParamsForSide(self, isLeft: bool) -> None:
        if self._parameterNode is None:
            return
        s = "Left" if isLeft else "Right"
        self._parameterNode.__setattr__(f"huMin{s}",        self.ui.huMinSpinBox.value)
        self._parameterNode.__setattr__(f"huMax{s}",        self.ui.huMaxSpinBox.value)
        self._parameterNode.__setattr__(f"autoSeed{s}",     self.ui.autoSeedCheckBox.isChecked())
        self._parameterNode.__setattr__(f"seedOffset{s}",   self.ui.seedOffsetSpinBox.value)
        self._parameterNode.__setattr__(f"radiusMargin{s}", self.ui.radiusMarginSpinBox.value)
        self._parameterNode.__setattr__(f"stoppingValue{s}",self.ui.stoppingValueSpinBox.value)
        self._parameterNode.__setattr__(f"speedSigma{s}",   self.ui.speedSigmaSpinBox.value)
        self._parameterNode.__setattr__(f"posteriorBoost{s}",self.ui.posteriorBoostSpinBox.value)
        self._parameterNode.__setattr__(f"showSeed{s}", self.ui.showSeedCheckBox.isChecked())
        mode = (1 if self.ui.rbSeedContralateral.isChecked()
                else 2 if self.ui.rbSeedModelBased.isChecked() else 0)
        self._parameterNode.__setattr__(f"seedMode{s}", mode)
        if isLeft:
            self._parameterNode.modelSeedNodeLeft  = self.ui.modelSeedSelector.currentNode()
        else:
            self._parameterNode.modelSeedNodeRight = self.ui.modelSeedSelector.currentNode()
        self._parameterNode.__setattr__(f"removeSatellites{s}", self.ui.removeSatellitesCheckBox.isChecked())
        self._parameterNode.__setattr__(f"satelliteDiam{s}",    self.ui.satelliteDiamSpinBox.value)

    def _loadParamsForSide(self, isLeft: bool) -> None:
        if self._parameterNode is None:
            return
        s = "Left" if isLeft else "Right"
        self.ui.huMinSpinBox.setValue(        self._parameterNode.__getattribute__(f"huMin{s}"))
        self.ui.huMaxSpinBox.setValue(        self._parameterNode.__getattribute__(f"huMax{s}"))
        autoSeed = self._parameterNode.__getattribute__(f"autoSeed{s}")
        self.ui.autoSeedCheckBox.setChecked(autoSeed)
        self.ui.seedOffsetSpinBox.setValue(   self._parameterNode.__getattribute__(f"seedOffset{s}"))
        self.ui.seedOffsetSpinBox.setEnabled(not autoSeed)
        self.ui.radiusMarginSpinBox.setValue( self._parameterNode.__getattribute__(f"radiusMargin{s}"))
        self.ui.stoppingValueSpinBox.setValue(self._parameterNode.__getattribute__(f"stoppingValue{s}"))
        self.ui.speedSigmaSpinBox.setValue(   self._parameterNode.__getattribute__(f"speedSigma{s}"))
        self.ui.posteriorBoostSpinBox.setValue(self._parameterNode.__getattribute__(f"posteriorBoost{s}"))
        self.ui.showSeedCheckBox.setChecked(self._parameterNode.__getattribute__(f"showSeed{s}"))
        # Restore seed-mode radio buttons without triggering onSeedModeChanged
        mode = self._parameterNode.__getattribute__(f"seedMode{s}")
        for rb in [self.ui.rbSeedManual, self.ui.rbSeedContralateral, self.ui.rbSeedModelBased]:
            rb.blockSignals(True)
        self.ui.rbSeedManual.setChecked(mode == 0)
        self.ui.rbSeedContralateral.setChecked(mode == 1)
        self.ui.rbSeedModelBased.setChecked(mode == 2)
        for rb in [self.ui.rbSeedManual, self.ui.rbSeedContralateral, self.ui.rbSeedModelBased]:
            rb.blockSignals(False)
        # Contralateral option only available when the other side is already segmented
        contra_node = (self._parameterNode.segmentationNodeRight
                       if isLeft else self._parameterNode.segmentationNodeLeft)
        self.ui.rbSeedContralateral.setEnabled(contra_node is not None)
        # Show/hide mode-specific rows
        is_contra = (mode == 1)
        is_model  = (mode == 2)
        self.ui.labelContraPosition.setVisible(is_contra)
        self.ui.positionContralateralWidget.setVisible(is_contra)
        self.ui.positionContralateralButton.setEnabled(
            is_contra and contra_node is not None
        )
        self.ui.labelModelSeed.setVisible(is_model)
        self.ui.modelSeedWidget.setVisible(is_model)
        # Restore model seed selector
        model_seed = getattr(self._parameterNode, f"modelSeedNode{s}", None)
        self.ui.modelSeedSelector.blockSignals(True)
        self.ui.modelSeedSelector.setCurrentNode(model_seed)
        self.ui.modelSeedSelector.blockSignals(False)
        self.ui.positionModelButton.setEnabled(model_seed is not None)
        self.ui.mirrorTemplateButton.setEnabled(model_seed is not None)
        remove_sat = self._parameterNode.__getattribute__(f"removeSatellites{s}")
        self.ui.removeSatellitesCheckBox.setChecked(remove_sat)
        self.ui.satelliteDiamSpinBox.setValue(       self._parameterNode.__getattribute__(f"satelliteDiam{s}"))
        self.ui.satelliteDiamSpinBox.setEnabled(remove_sat)

    def _copyParamsToSide(self, fromLeft: bool) -> None:
        """Kopiert alle Segmentierungsparameter von einer Seite zur anderen."""
        if self._parameterNode is None:
            return
        src = "Left" if fromLeft else "Right"
        dst = "Right" if fromLeft else "Left"
        for param in ["huMin", "huMax", "autoSeed", "seedOffset", 
                      "radiusMargin", "stoppingValue", "speedSigma", "posteriorBoost",
                      "showSeed", "seedMode", "removeSatellites", "satelliteDiam"]:
            val = self._parameterNode.__getattribute__(f"{param}{src}")
            self._parameterNode.__setattr__(f"{param}{dst}", val)

    def onCurveChanged(self, node) -> None:
        isLeft = self._parameterNode.sideIsLeft if self._parameterNode else True
        self._curveNodes[isLeft] = node
        self.ui.createSurfaceButton.setEnabled(node is not None)

    def onPlaneModelChanged(self, node) -> None:
        if self._parameterNode is None:
            return
        if self._parameterNode.sideIsLeft:
            self._parameterNode.planeModelLeft = node
        else:
            self._parameterNode.planeModelRight = node

    def onCreateCurveButton(self) -> None:
        isLeft = self._parameterNode.sideIsLeft if self._parameterNode else True
        side_suffix = "L" if isLeft else "R"

        curve_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLMarkupsClosedCurveNode")
        curve_node.SetName(f"OrbitalRim_{side_suffix}")
        disp = curve_node.GetDisplayNode()
        if disp:
            disp.SetColor(0.0, 1.0, 0.0)
            disp.SetGlyphScale(3.0)

        self._curveNodes[isLeft] = curve_node
        self._placeInFolder(curve_node, isLeft)
        self.ui.curveSelector.blockSignals(True)
        self.ui.curveSelector.setCurrentNode(curve_node)
        self.ui.curveSelector.blockSignals(False)
        self.ui.createSurfaceButton.setEnabled(True)

        # Zeichen-Modus aktivieren
        selectionNode = slicer.app.applicationLogic().GetSelectionNode()
        selectionNode.SetActivePlaceNodeID(curve_node.GetID())
        selectionNode.SetActivePlaceNodeClassName("vtkMRMLMarkupsClosedCurveNode")
        interactionNode = slicer.app.applicationLogic().GetInteractionNode()
        interactionNode.SetCurrentInteractionMode(
            slicer.vtkMRMLInteractionNode.Place
        )

    def onToggleVolumeRendering(self, checked: bool) -> None:
        if self._vrDisplayNode is not None:
            self._vrDisplayNode.SetVisibility(checked)
            self._vrDisplayNode.SetVisibility3D(checked)
        self.ui.toggleVolumeRenderingButton.setText(_("3D on") if not checked else _("3D off"))

    def onCTVolumeChanged(self, volume_node) -> None:
        if self._parameterNode is not None:
            self._parameterNode.ctVolume = volume_node
        if volume_node is None:
            self.ui.currentSpacingLabel.setText("—")
            return

        from PSOILib import helperfunctions
        helperfunctions.hideAllVolumeRenderingNodes()
        self._vrDisplayNode = None
        self._applyVolumeRendering(volume_node)
        self._centerViewAnterior(volume_node)
        slicer.util.setSliceViewerLayers(background=volume_node)

        # Populate slice thickness row
        spacing = volume_node.GetSpacing()  # (x, y, z) in mm
        min_spacing = min(spacing)
        self.ui.currentSpacingLabel.setText(f"Current: {min_spacing:.2f} mm")
        self.ui.targetSpacingSpinBox.setValue(min_spacing)

        # Warn if spacing differs from other parameter nodes' CTs
        self._checkSpacingConsistency(volume_node)

        # determine minimal and maximal HU Values
        volumeArray = slicer.util.arrayFromVolume(volume_node).flatten()

        hu_min = np.min(volumeArray)
        hu_max = np.max(volumeArray)

        self.ui.huMinSpinBox.minimum = hu_min
        self.ui.huMaxSpinBox.minimum = hu_min
        self.ui.huMinSpinBox.maximum = hu_max
        self.ui.huMaxSpinBox.maximum = hu_max

    def _checkSpacingConsistency(self, new_volume_node) -> None:
        """Warn if new CT spacing differs from other parameter nodes' CT spacings."""
        new_spacing = min(new_volume_node.GetSpacing())
        mismatches = []
        for mrml_node in self.logic.getAllParameterNodes():
            pn = OrbitalVolumeWorkflowModuleParameterNode(mrml_node)
            other_vol = pn.ctVolume
            if other_vol is None or other_vol.GetID() == new_volume_node.GetID():
                continue
            other_spacing = min(other_vol.GetSpacing())
            if abs(other_spacing - new_spacing) > 0.01:
                mismatches.append(
                    f"  • {mrml_node.GetName()}: {other_spacing:.2f} mm  ≠  {new_spacing:.2f} mm"
                )
        if mismatches:
            msg = (
                "The slice thickness of the newly selected CT differs from other parameter sets:\n\n"
                + "\n".join(mismatches)
                + "\n\nConsider resampling all volumes to the same isotropic spacing "
                "for better comparability."
            )
            slicer.util.warningDisplay(msg, windowTitle="Slice thickness mismatch")

    def onResampleCTButton(self, clicked: bool = False) -> None:
        volume_node = self.ui.ctVolumeSelector.currentNode()
        if volume_node is None:
            slicer.util.warningDisplay("No CT volume selected.", windowTitle="Resample CT")
            return
        target_spacing = self.ui.targetSpacingSpinBox.value
        current_spacing = min(volume_node.GetSpacing())
        if abs(target_spacing - current_spacing) < 0.005:
            slicer.util.infoDisplay(
                f"CT is already at {current_spacing:.2f} mm spacing — no resampling needed.",
                windowTitle="Resample CT",
            )
            return
        resampled = self.logic.resampleVolume(volume_node, target_spacing)
        if resampled is None:
            slicer.util.errorDisplay("Resampling failed.", windowTitle="Resample CT")
            return
        # Switch CT selector to the new volume
        self.ui.ctVolumeSelector.setCurrentNode(resampled)

    def onHUShiftChanged(self, shift_hu: float) -> None:
        """
            Get's called when the user changes the HU-Shift. Shifts the 
            Preset's values by the amount given
        """

        # get Display Node and Properties
        if self._vrDisplayNode is None:
            print("No vrDisplayNode set")
            return
        
        volPropNode = self._vrDisplayNode.GetVolumePropertyNode()
        
        if volPropNode is None:
            return
        
        # get Preset values
        vrLogic = slicer.modules.volumerendering.logic()
        presetNode = vrLogic.GetPresetByName("CT-Bone")
        
        if presetNode is None:
            return
        
        presetTransferFunction = presetNode.GetScalarOpacity()
        gradientOpacityTransferFunction = volPropNode.GetScalarOpacity()
        
        values = [0,0,0,0]
        for i in range(gradientOpacityTransferFunction.GetSize()):
            presetTransferFunction.GetNodeValue(i, values)
            values[0] += shift_hu
            gradientOpacityTransferFunction.SetNodeValue(i, values)

        """"
        # Preset frisch kopieren, dann Shift anwenden
        volPropNode.Copy(presetNode)
        volProp = volPropNode.GetVolumeProperty()

        opFn = volProp.GetScalarOpacity()
        n = opFn.GetSize()
        op_nodes = []
        for i in range(n):
            v = [0.0] * 4
            opFn.GetNodeValue(i, v)
            op_nodes.append(v[:])
        opFn.RemoveAllPoints()
        for v in op_nodes:
            opFn.AddPoint(v[0] + shift_hu, v[1], v[2], v[3])

        colFn = volProp.GetRGBTransferFunction()
        n = colFn.GetSize()
        col_nodes = []
        for i in range(n):
            v = [0.0] * 6
            colFn.GetNodeValue(i, v)
            col_nodes.append(v[:])
        colFn.RemoveAllPoints()
        for v in col_nodes:
            colFn.AddRGBPoint(v[0] + shift_hu, v[1], v[2], v[3], v[4], v[5])

        volPropNode.Modified()
        """

    def _applyVolumeRendering(self, volume_node) -> None:
        from PSOILib import helperfunctions

        self._vrDisplayNode = helperfunctions.showVolumeRendering(volume_node, preset="CT-Bone")

        # DisplayNode für den HU-Shift-Mechanismus referenzieren
        # vrLogic = slicer.modules.volumerendering.logic()
        # self._vrDisplayNode = vrLogic.CreateDefaultVolumeRenderingNodes(volume_node)

        # Toggle-Button in den "ein"-Zustand bringen (ohne erneutes Signal)
        self.ui.toggleVolumeRenderingButton.blockSignals(True)
        self.ui.toggleVolumeRenderingButton.setChecked(True)
        self.ui.toggleVolumeRenderingButton.setText(_("3D off"))
        self.ui.toggleVolumeRenderingButton.blockSignals(False)

        # Aktuellen HU-Shift anwenden (überschreibt ggf. das frisch gesetzte Preset)
        self.onHUShiftChanged(self.ui.huShiftSlider.value)

    def _centerViewAnterior(self, volume_node) -> None:
        layoutManager = slicer.app.layoutManager()
        if layoutManager is None:
            return
        threeDWidget = layoutManager.threeDWidget(0)
        if threeDWidget is None:
            return
        threeDView = threeDWidget.threeDView()

        bounds = [0.0] * 6
        volume_node.GetRASBounds(bounds)
        cx = (bounds[0] + bounds[1]) / 2
        cy = (bounds[2] + bounds[3]) / 2
        cz = (bounds[4] + bounds[5]) / 2
        extent = max(bounds[1]-bounds[0], bounds[3]-bounds[2], bounds[5]-bounds[4])

        viewNode = threeDView.mrmlViewNode()
        cameraNode = slicer.modules.cameras.logic().GetViewActiveCameraNode(viewNode)
        if cameraNode is None:
            return

        # Von anterior betrachten: Kamera auf +Y (RAS: A = anterior = +Y)
        cameraNode.SetFocalPoint(cx, cy, cz)
        cameraNode.SetPosition(cx, cy + extent * 1.5, cz)
        cameraNode.SetViewUp(0.0, 0.0, 1.0)
        cameraNode.ResetClippingRange()
        threeDView.forceRender()

    # ------------------------------------------------------------------
    # Signalhandler – Workflow
    # ------------------------------------------------------------------

    def onStepsToolboxCurrentChanged(self, currentId: int) -> None:
        if self._parameterNode:
            self._parameterNode.step = currentId

    def onAutoSeedToggled(self, checked: bool) -> None:
        self.ui.seedOffsetSpinBox.setEnabled(not checked)

    def onSeedModeChanged(self) -> None:
        is_contra = self.ui.rbSeedContralateral.isChecked()
        is_model  = self.ui.rbSeedModelBased.isChecked()
        # For region-seed modes the FM starts from a large pre-computed volume,
        # so a much lower stopping value is sufficient.
        self.ui.stoppingValueSpinBox.setValue(15.0 if (is_contra or is_model) else 25.0)
        # Show / hide mode-specific rows
        self.ui.labelContraPosition.setVisible(is_contra)
        
        self.ui.positionContralateralWidget.setVisible(is_contra)
        self.ui.labelModelSeed.setVisible(is_model)
        self.ui.modelSeedWidget.setVisible(is_model)
        # Enable "Position Mirrored" only when the contralateral segmentation exists
        if is_contra and self._parameterNode is not None:
            isLeft = self._parameterNode.sideIsLeft
            contra_node = (self._parameterNode.segmentationNodeRight if isLeft
                           else self._parameterNode.segmentationNodeLeft)
            self.ui.positionContralateralButton.setEnabled(contra_node is not None)
        else:
            self.ui.positionContralateralButton.setEnabled(False)

    def onModelSeedChanged(self, node) -> None:
        self.ui.positionModelButton.setEnabled(node is not None)
        self.ui.mirrorTemplateButton.setEnabled(node is not None)
        if self._parameterNode is None:
            return
        isLeft = self._parameterNode.sideIsLeft
        if isLeft:
            self._parameterNode.modelSeedNodeLeft  = node
        else:
            self._parameterNode.modelSeedNodeRight = node

    def onMirrorTemplateToggled(self, checked: bool) -> None:
        """Spiegelt das Template-Volumen entlang der Mediansagittalebene (X=0) und härtet
        den Spiegelungs-Transform sofort, damit die Normalen korrekt bleiben.

        Da eine Reflexion involutorisch ist (M² = I), ist das Vorgehen für Ein- und
        Ausschalten identisch: den Mirror-Transform anwenden und härten.  T_centering
        wird danach mit dem neuen Schwerpunkt des gehärteten Volumens aktualisiert.
        """
        if self._parameterNode is None:
            return
        isLeft      = self._parameterNode.sideIsLeft
        seg_node    = self.ui.modelSeedSelector.currentNode()
        if seg_node is None:
            self.ui.mirrorTemplateButton.blockSignals(True)
            self.ui.mirrorTemplateButton.setChecked(False)
            self.ui.mirrorTemplateButton.blockSignals(False)
            return

        side_suffix    = "L" if isLeft else "R"
        centering_name = f"ModelSeedCentering_{side_suffix}"
        tf_attr        = "modelSeedTransformLeft" if isLeft else "modelSeedTransformRight"
        transform_node = getattr(self._parameterNode, tf_attr, None)

        centering_node = slicer.mrmlScene.GetFirstNodeByName(centering_name)
        if centering_node is None:
            slicer.util.warningDisplay(_("Please click 'Position Model' first."))
            self.ui.mirrorTemplateButton.blockSignals(True)
            self.ui.mirrorTemplateButton.setChecked(False)
            self.ui.mirrorTemplateButton.blockSignals(False)
            return

        # ── Segment aus der Positionierungs-Kette lösen ────────────────────
        seg_node.SetAndObserveTransformNodeID(None)

        # ── Reflexionsmatrix X → −X anwenden und sofort härten ─────────────
        tmp_mirror = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLinearTransformNode", "_TmpMirrorHarden"
        )
        mat_m = vtk.vtkMatrix4x4()
        mat_m.Identity()
        mat_m.SetElement(0, 0, -1.0)
        tmp_mirror.SetMatrixTransformToParent(mat_m)
        seg_node.SetAndObserveTransformNodeID(tmp_mirror.GetID())
        slicer.modules.transforms.logic().hardenTransform(seg_node)
        slicer.mrmlScene.RemoveNode(tmp_mirror)

        # ── Neuen Schwerpunkt nach dem Härten bestimmen ─────────────────────
        seg_bounds = [0.0] * 6
        seg_node.GetRASBounds(seg_bounds)
        if seg_bounds[0] < seg_bounds[1]:
            seg_center = np.array([
                (seg_bounds[0] + seg_bounds[1]) / 2,
                (seg_bounds[2] + seg_bounds[3]) / 2,
                (seg_bounds[4] + seg_bounds[5]) / 2,
            ])
        else:
            seg_center = np.zeros(3)

        # ── T_centering mit neuem Schwerpunkt aktualisieren ─────────────────
        mat_c = vtk.vtkMatrix4x4()
        mat_c.Identity()
        mat_c.SetElement(0, 3, float(-seg_center[0]))
        mat_c.SetElement(1, 3, float(-seg_center[1]))
        mat_c.SetElement(2, 3, float(-seg_center[2]))
        centering_node.SetMatrixTransformToParent(mat_c)

        # ── Hierarchie wiederherstellen: seg → T_centering → T_main ─────────
        centering_node.SetAndObserveTransformNodeID(
            transform_node.GetID() if transform_node else None
        )
        seg_node.SetAndObserveTransformNodeID(centering_node.GetID())

    def onLoadTemplateButton(self) -> None:
        path = qt.QFileDialog.getOpenFileName(
            slicer.util.mainWindow(),
            _("Load template segmentation"),
            "",
            "Segmentation files (*.seg.nrrd *.nrrd *.nii *.nii.gz);;All files (*)",
        )
        if not path:
            return
        raw = slicer.util.loadSegmentation(path)
        if raw is None:
            slicer.util.errorDisplay(_("Could not load segmentation from file."))
            return
        # Kopie anlegen, damit das Original unverändert bleibt und derselbe
        # Template-File mehrfach (für verschiedene Seiten / Parameter-Nodes) verwendet
        # werden kann ohne Konflikte durch Transforms oder Umbenennung.
        isLeft = self._parameterNode is not None and self._parameterNode.sideIsLeft
        side_suffix = "L" if isLeft else "R"
        pn_name = (self._parameterNode.parameterNode.GetName()
                   if self._parameterNode is not None else "Template")
        copy_node = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLSegmentationNode", f"{pn_name}_Template_{side_suffix}"
        )
        copy_node.GetSegmentation().DeepCopy(raw.GetSegmentation())
        copy_node.CreateDefaultDisplayNodes()
        slicer.mrmlScene.RemoveNode(raw)
        self.ui.modelSeedSelector.setCurrentNode(copy_node)

    def onPositionModelButton(self) -> None:
        if self._parameterNode is None:
            return
        isLeft      = self._parameterNode.sideIsLeft
        seg_node    = self.ui.modelSeedSelector.currentNode()
        volume_node = self.ui.ctVolumeSelector.currentNode()
        if seg_node is None:
            slicer.util.warningDisplay(_("Please select a template segmentation first."))
            return

        side_suffix = "L" if isLeft else "R"

        # ── Haupttransform (sichtbar, interaktiv) ──────────────────────────
        tf_attr = "modelSeedTransformLeft" if isLeft else "modelSeedTransformRight"
        existing_tf = getattr(self._parameterNode, tf_attr, None)
        if existing_tf is not None and slicer.mrmlScene.IsNodePresent(existing_tf):
            transform_node = existing_tf
        else:
            transform_node = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLLinearTransformNode", f"ModelSeedTransform_{side_suffix}"
            )
            transform_node.CreateDefaultDisplayNodes()
            setattr(self._parameterNode, tf_attr, transform_node)
        self._placeInFolder(transform_node, isLeft)

        # ── Zentrierungs-Transform (unsichtbar, Kind von T_main) ────────────
        centering_name = f"ModelSeedCentering_{side_suffix}"
        centering_node = slicer.mrmlScene.GetFirstNodeByName(centering_name)
        if centering_node is None:
            centering_node = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLLinearTransformNode", centering_name
            )
            centering_node.CreateDefaultDisplayNodes()
        self._placeInFolder(centering_node, isLeft)

        # ── Zielposition: Orbital-Ebenen-Centroid + 20 mm posterior ─────────
        plane = (self._parameterNode.planeModelLeft if isLeft
                 else self._parameterNode.planeModelRight)
        if plane is not None:
            centroid, normal = self.logic._getPlaneFromModel(plane)
            if volume_node is not None:
                posterior = self.logic._ensurePosteriorDirection(normal, centroid, volume_node)
            else:
                posterior = normal
            target = centroid + 20.0 * posterior
        elif volume_node is not None:
            bounds = [0.0] * 6
            volume_node.GetRASBounds(bounds)
            centroid = np.array([(bounds[0]+bounds[1])/2,
                                  (bounds[2]+bounds[3])/2,
                                  (bounds[4]+bounds[5])/2])
            target = centroid
        else:
            target = np.zeros(3)

        # ── Natürlichen Mittelpunkt der Template-Segmentierung bestimmen ─────
        # Zuerst alle Transforms lösen, damit GetRASBounds unkompensierte Koordinaten liefert.
        seg_node.SetAndObserveTransformNodeID(None)
        seg_bounds = [0.0] * 6
        seg_node.GetRASBounds(seg_bounds)
        if seg_bounds[0] < seg_bounds[1]:   # gültige Bounds
            seg_center = np.array([
                (seg_bounds[0] + seg_bounds[1]) / 2,
                (seg_bounds[2] + seg_bounds[3]) / 2,
                (seg_bounds[4] + seg_bounds[5]) / 2,
            ])
        else:
            seg_center = np.zeros(3)

        # ── Zwei-Transform-Hierarchie aufbauen ───────────────────────────────
        # T_centering: [I | -seg_center]  →  verschiebt Template-Mittelpunkt auf Ursprung
        mat_c = vtk.vtkMatrix4x4()
        mat_c.Identity()
        mat_c.SetElement(0, 3, float(-seg_center[0]))
        mat_c.SetElement(1, 3, float(-seg_center[1]))
        mat_c.SetElement(2, 3, float(-seg_center[2]))
        centering_node.SetMatrixTransformToParent(mat_c)

        # T_main: [I | target]  →  Translations-Vektor = Zielposition = Center of Transformation
        mat_m = vtk.vtkMatrix4x4()
        mat_m.Identity()
        mat_m.SetElement(0, 3, float(target[0]))
        mat_m.SetElement(1, 3, float(target[1]))
        mat_m.SetElement(2, 3, float(target[2]))
        transform_node.SetMatrixTransformToParent(mat_m)

        # Hierarchie: seg_node → T_centering → T_main
        centering_node.SetAndObserveTransformNodeID(transform_node.GetID())
        seg_node.SetAndObserveTransformNodeID(centering_node.GetID())

        # T_centering unsichtbar (keine Handles)
        c_disp = centering_node.GetDisplayNode()
        if c_disp:
            c_disp.SetEditorVisibility(False)

        # T_main: Handles nur in 2D-Slice-Views
        disp = transform_node.GetDisplayNode()
        disp.SetEditorVisibility(True)
        # 3D-Handles explizit deaktivieren (Slicer-Default wäre True)
        disp.SetEditorTranslationEnabled(False)
        disp.SetEditorRotationEnabled(False)
        disp.SetEditorScalingEnabled(False)
        # 2D-Slice-Handles aktivieren
        disp.SetEditorTranslationSliceEnabled(True)
        disp.SetEditorRotationSliceEnabled(True)
        disp.SetEditorScalingSliceEnabled(True)
        disp.SetScaleHandleComponentVisibilitySlice([True, True, True, False])
        disp.SetTranslationHandleComponentVisibilitySlice([True, True, True, True])

        slicer.app.layoutManager().threeDWidget(0).threeDView().resetFocalPoint()

    def onPositionContralateralButton(self) -> None:
        """Spiegelt das kontralaterale Segment und positioniert die Spiegelung
        mit interaktiven 2D-Handles – analog zu onPositionModelButton().

        Das erzeugte 'ContraPositioned_{s}'-Segment wird beim nächsten Klick
        auf 'Segment Volume' (Modus 1) statt der automatischen Spiegelung als
        FM-Seed verwendet.
        """
        if self._parameterNode is None:
            return
        isLeft  = self._parameterNode.sideIsLeft
        s       = "Left" if isLeft else "Right"
        side_suffix = "L" if isLeft else "R"

        contra_node = (self._parameterNode.segmentationNodeRight if isLeft
                       else self._parameterNode.segmentationNodeLeft)
        if contra_node is None:
            slicer.util.warningDisplay(_("Bitte zuerst die Gegenseite segmentieren."))
            return

        # Gegenseiten-Segment-ID ermitteln
        contra_side  = not isLeft
        contra_seg_id = self._segIds.get(contra_side)
        if contra_seg_id is None:
            seg = contra_node.GetSegmentation()
            for i in range(seg.GetNumberOfSegments()):
                if seg.GetNthSegment(i).GetName() == "IntraorbitalVolume":
                    contra_seg_id = seg.GetNthSegmentID(i)
                    break
        if contra_seg_id is None:
            slicer.util.warningDisplay(
                _("Kein IntraorbitalVolume-Segment auf der Gegenseite gefunden.")
            )
            return

        volume_node = self.ui.ctVolumeSelector.currentNode()

        import SimpleITK as sitk
        import sitkUtils

        # Gegenseiten-Segment in Labelmap exportieren und spiegeln (L-R)
        seg_id_arr = vtk.vtkStringArray()
        seg_id_arr.InsertNextValue(contra_seg_id)
        tmp_lm = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode", "_TmpContraLM_Pos"
        )
        try:
            slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                contra_node, seg_id_arr, tmp_lm, volume_node
            )
            contra_sitk = sitkUtils.PullVolumeFromSlicer(tmp_lm)
        finally:
            slicer.mrmlScene.RemoveNode(tmp_lm)

        contra_arr = sitk.GetArrayFromImage(contra_sitk)
        mirrored = np.ascontiguousarray(np.flip(contra_arr, axis=2))
        mirrored_sitk = sitk.GetImageFromArray(mirrored)
        mirrored_sitk.CopyInformation(contra_sitk)

        # Positionier-Node anlegen oder wiederverwenden
        pos_node_attr = f"contraPositionedNode{s}"
        existing_pos  = getattr(self._parameterNode, pos_node_attr, None)
        if existing_pos is not None and slicer.mrmlScene.IsNodePresent(existing_pos):
            existing_pos.SetAndObserveTransformNodeID(None)
            existing_pos.GetSegmentation().RemoveAllSegments()
            pos_node = existing_pos
        else:
            pos_node = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLSegmentationNode", f"ContraPositioned_{side_suffix}"
            )
            pos_node.CreateDefaultDisplayNodes()
            if volume_node:
                pos_node.SetReferenceImageGeometryParameterFromVolumeNode(volume_node)
            setattr(self._parameterNode, pos_node_attr, pos_node)
        self._placeInFolder(pos_node, isLeft)

        # Gespiegelte Maske importieren
        lm_tmp = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode", "_TmpMirroredLM"
        )
        sitkUtils.PushVolumeToSlicer(mirrored_sitk, lm_tmp)
        try:
            slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
                lm_tmp, pos_node
            )
        finally:
            slicer.mrmlScene.RemoveNode(lm_tmp)

        # Darstellung: blau, halbtransparent
        disp_seg = pos_node.GetDisplayNode()
        if disp_seg:
            disp_seg.SetOpacity3D(0.4)
            disp_seg.SetOpacity2DFill(0.3)
        pos_seg = pos_node.GetSegmentation()
        if pos_seg.GetNumberOfSegments() > 0:
            pos_seg.GetNthSegment(0).SetColor(0.2, 0.6, 1.0)

        # ── Haupt-Transform (interaktiv) ──────────────────────────────────
        tf_attr    = f"contraPositionedTransform{s}"
        existing_tf = getattr(self._parameterNode, tf_attr, None)
        if existing_tf is not None and slicer.mrmlScene.IsNodePresent(existing_tf):
            transform_node = existing_tf
        else:
            transform_node = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLLinearTransformNode", f"ContraPositionedTransform_{side_suffix}"
            )
            transform_node.CreateDefaultDisplayNodes()
            setattr(self._parameterNode, tf_attr, transform_node)
        self._placeInFolder(transform_node, isLeft)

        # ── Zentrierungs-Transform (unsichtbar) ───────────────────────────
        centering_name = f"ContraPositionedCentering_{side_suffix}"
        centering_node = slicer.mrmlScene.GetFirstNodeByName(centering_name)
        if centering_node is None:
            centering_node = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLLinearTransformNode", centering_name
            )
            centering_node.CreateDefaultDisplayNodes()
        self._placeInFolder(centering_node, isLeft)

        # ── Zielposition: Orbital-Ebenen-Centroid + 20 mm posterior ───────
        plane = (self._parameterNode.planeModelLeft if isLeft
                 else self._parameterNode.planeModelRight)
        if plane is not None:
            centroid, normal = self.logic._getPlaneFromModel(plane)
            if volume_node is not None:
                posterior = self.logic._ensurePosteriorDirection(normal, centroid, volume_node)
            else:
                posterior = normal
            target = centroid + 20.0 * posterior
        elif volume_node is not None:
            bounds = [0.0] * 6
            volume_node.GetRASBounds(bounds)
            target = np.array([
                (bounds[0]+bounds[1])/2,
                (bounds[2]+bounds[3])/2,
                (bounds[4]+bounds[5])/2,
            ])
        else:
            target = np.zeros(3)

        # Mittelpunkt des gespiegelten Segments
        seg_bounds = [0.0] * 6
        pos_node.GetRASBounds(seg_bounds)
        if seg_bounds[0] < seg_bounds[1]:
            seg_center = np.array([
                (seg_bounds[0]+seg_bounds[1])/2,
                (seg_bounds[2]+seg_bounds[3])/2,
                (seg_bounds[4]+seg_bounds[5])/2,
            ])
        else:
            seg_center = np.zeros(3)

        # ── Zwei-Transform-Hierarchie ──────────────────────────────────────
        mat_c = vtk.vtkMatrix4x4()
        mat_c.Identity()
        mat_c.SetElement(0, 3, float(-seg_center[0]))
        mat_c.SetElement(1, 3, float(-seg_center[1]))
        mat_c.SetElement(2, 3, float(-seg_center[2]))
        centering_node.SetMatrixTransformToParent(mat_c)

        mat_m = vtk.vtkMatrix4x4()
        mat_m.Identity()
        mat_m.SetElement(0, 3, float(target[0]))
        mat_m.SetElement(1, 3, float(target[1]))
        mat_m.SetElement(2, 3, float(target[2]))
        transform_node.SetMatrixTransformToParent(mat_m)

        centering_node.SetAndObserveTransformNodeID(transform_node.GetID())
        pos_node.SetAndObserveTransformNodeID(centering_node.GetID())

        # T_centering unsichtbar (keine Handles)
        c_disp = centering_node.GetDisplayNode()
        if c_disp:
            c_disp.SetEditorVisibility(False)

        # T_main: Handles nur in 2D-Slice-Views
        disp = transform_node.GetDisplayNode()
        disp.SetEditorVisibility(True)
        disp.SetEditorTranslationEnabled(False)
        disp.SetEditorRotationEnabled(False)
        disp.SetEditorScalingEnabled(False)
        disp.SetEditorTranslationSliceEnabled(True)
        disp.SetEditorRotationSliceEnabled(True)
        disp.SetEditorScalingSliceEnabled(True)
        disp.SetScaleHandleComponentVisibilitySlice([True, True, True, False])
        disp.SetTranslationHandleComponentVisibilitySlice([True, True, True, True])

        slicer.app.layoutManager().threeDWidget(0).threeDView().resetFocalPoint()

    def onExportResultsButton(self) -> None:
        """Exportiert Segmentierungsparameter und -ergebnisse beider Seiten als Excel-Datei
        in das Verzeichnis des CT-Volumes (results_orbital_volume.xlsx)."""
        import os

        if self._parameterNode is None:
            slicer.util.warningDisplay(_("No parameter node active."))
            return

        try:
            import openpyxl
        except ImportError:
            slicer.util.pip_install("openpyxl")
            import openpyxl

        

        # Zielverzeichnis aus dem CT-Volume ableiten
        volume_node = self._parameterNode.ctVolume
        ct_dir = None
        if volume_node:
            sn = volume_node.GetStorageNode()
            if sn and sn.GetFileName():
                ct_dir = os.path.dirname(sn.GetFileName())
        if not ct_dir:
            ct_dir = qt.QFileDialog.getExistingDirectory(
                slicer.util.mainWindow(), _("Select output directory")
            )
            if not ct_dir:
                return

        excel_path = os.path.join(ct_dir, "results_orbital_volume.xlsx")

        # Collect all module PNs now (needed for deletion logic below too)
        all_pn_nodes = []
        _col = slicer.mrmlScene.GetNodesByClass("vtkMRMLScriptedModuleNode")
        _col.InitTraversal()
        _node = _col.GetNextItemAsObject()
        while _node:
            if _node.GetAttribute("ModuleName") == self.moduleName:
                all_pn_nodes.append(_node)
            _node = _col.GetNextItemAsObject()
        export_pn_names = {(n.GetName() or "–") for n in all_pn_nodes}

        # Existierende Datei laden oder neu erstellen
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            # Vorhandene Zeilen für alle zu exportierenden PNs entfernen (Update-Semantik)
            rows_to_delete = [
                row_idx for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2)
                if row[0] in export_pn_names  # Spalte "Parameter-Node"
            ]
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Intraorbital Volume"
            headers = [
                "Parameter-Node", "CT-Volume", "Seite", "Seed-Modus",
                "Segmentierungsmethode",
                "HU Min", "HU Max",
                "Stopping Value (FM Limit)", "Speed Sigma (σ)",
                "Posteriorer Boost", "Satelliten entfernen",
                "Min. Satelliten-Ø (mm)", "Intraorbitalvolumen (ml)",
                "Gegenseite / Seite",
            ]
            ws.append(headers)
            # Spaltenbreiten
            for col, width in enumerate([28, 22, 8, 24, 22, 8, 8, 24, 16, 18, 22, 22, 24, 18], start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        seed_mode_labels = {0: "Manuell", 1: "Gegenseite gespiegelt", 2: "Modellbasiert"}

        for raw_pn in all_pn_nodes:
            pn = OrbitalVolumeWorkflowModuleParameterNode(raw_pn)
            pn_name = raw_pn.GetName() if raw_pn.GetName() else "–"

            # CT volume and output directory
            pn_vol_node = pn.ctVolume
            pn_ct_name = pn_vol_node.GetName() if pn_vol_node else "–"

            # Compute volumes for both sides directly from the segmentation nodes.
            # Always recompute (never use the in-memory cache) so that volumes are
            # consistent regardless of which PN was last active.
            vols = {}
            for isLeft in [True, False]:
                seg_node = (pn.segmentationNodeLeft if isLeft else pn.segmentationNodeRight)
                if seg_node is None:
                    vols[isLeft] = None
                    continue
                seg = seg_node.GetSegmentation()
                seg_id = None
                for i in range(seg.GetNumberOfSegments()):
                    if seg.GetNthSegment(i).GetName() == "IntraorbitalVolume":
                        seg_id = seg.GetNthSegmentID(i)
                        break
                if seg_id is None:
                    vols[isLeft] = None
                else:
                    vol, _vox = self._calculateVolumeFromSegment(seg_node, seg_id)
                    vols[isLeft] = vol

            for isLeft in [True, False]:
                seg_node = (pn.segmentationNodeLeft if isLeft else pn.segmentationNodeRight)
                if seg_node is None:
                    continue

                s = "Left" if isLeft else "Right"
                vol_ml    = vols[isLeft]
                other_vol = vols[not isLeft]

                if vol_ml and other_vol and vol_ml > 0:
                    ratio = vol_ml / other_vol
                else:
                    ratio = None

                method_raw = pn.__getattribute__(f"segMethod{s}")
                method_label = {"fastmarching": "Fast Marching",
                                "threshold":    "Threshold"}.get(method_raw, "–")
                row = [
                    pn_name,
                    pn_ct_name,
                    "Links" if isLeft else "Rechts",
                    seed_mode_labels.get(pn.__getattribute__(f"seedMode{s}"), "–"),
                    method_label,
                    pn.__getattribute__(f"huMin{s}"),
                    pn.__getattribute__(f"huMax{s}"),
                    pn.__getattribute__(f"stoppingValue{s}"),
                    pn.__getattribute__(f"speedSigma{s}"),
                    pn.__getattribute__(f"posteriorBoost{s}"),
                    "Ja" if pn.__getattribute__(f"removeSatellites{s}") else "Nein",
                    pn.__getattribute__(f"satelliteDiam{s}"),
                    round(vol_ml, 3) if vol_ml is not None else "–",
                    ratio if ratio is not None else "–",
                ]
                ws.append(row)

                if ratio is not None:
                    ws.cell(row=ws.max_row, column=13).number_format = "0.0%"

        from openpyxl.styles import Font, Color, Border, Side

        # Allen Zellen die Schriftfarbe "auto" zuweisen, damit es auch im Dark mode
        # richtig dargestellt wird
        cell_font = Font(name="Liberations Sans", color=Color(auto=True))
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font

        # Titelzeile fett darstellen
        header_font = Font(name="Liberations Sans", color=Color(auto=True), bold=True)
        header_border = Border(bottom=Side(border_style="thin",color=Color(auto=True)))   

        for cell in ws[1]:
            cell.font = header_font
            cell.border = header_border

        # Ergebnis epeichern
        wb.save(excel_path)
        slicer.util.infoDisplay(
            _("Results saved to:\n{path}").format(path=excel_path)
        )

    def onCreateSurfaceButton(self) -> None:
        with slicer.util.tryWithErrorDisplay(_("Error creating the orbital plane."), waitCursor=True):
            curve_node = self.ui.curveSelector.currentNode()
            if curve_node is None:
                raise ValueError(_("Please select a closed curve (Closed Curve)."))

            isLeft = self._parameterNode.sideIsLeft
            color = (0.2, 0.7, 1.0) if isLeft else (1.0, 0.55, 0.1)

            model_node = self.logic.createOrbitalSurface(
                curve_node=curve_node,
                method=self.ui.methodComboBox.currentText,
                subdivision_distance=self.ui.subdivisionDistanceSpinBox.value,
                smooth_iterations=self.ui.smoothIterationsSpinBox.value,
                color=color,
            )

            n_tris = model_node.GetPolyData().GetNumberOfCells()
            n_pts  = model_node.GetPolyData().GetNumberOfPoints()
            text = (
                _("<b>Created:</b> {name}<br>Triangles: {tris} &nbsp;|&nbsp; Points: {pts}").format(
                    name=model_node.GetName(), tris=n_tris, pts=n_pts
                )
            )
            self._surfaceTexts[isLeft] = text
            self.ui.surfaceResultLabel.setText(text)

            if isLeft:
                self._parameterNode.planeModelLeft = model_node
            else:
                self._parameterNode.planeModelRight = model_node
            self._placeInFolder(model_node, isLeft)

            self.ui.planeModelSelector.blockSignals(True)
            self.ui.planeModelSelector.setCurrentNode(model_node)
            self.ui.planeModelSelector.blockSignals(False)

            self._nextStep()

    def onFastMarchingButton(self) -> None:
        self._runSegmentation('fastmarching')

    def onThresholdButton(self) -> None:
        self._runSegmentation('threshold')

    def _runSegmentation(self, method: str) -> None:
        with slicer.util.tryWithErrorDisplay(_("Error during volume segmentation."), waitCursor=True):
            volume_node = self.ui.ctVolumeSelector.currentNode()
            plane_model = self.ui.planeModelSelector.currentNode()

            if volume_node is None:
                raise ValueError(_("Please select a CT volume."))
            if plane_model is None:
                raise ValueError(_("Please select an orbital plane (mesh)."))

            isLeft = self._parameterNode.sideIsLeft
            self._saveParamsForSide(isLeft)
            existing_seg  = (self._parameterNode.segmentationNodeLeft
                             if isLeft else self._parameterNode.segmentationNodeRight)
            existing_seed = (self._parameterNode.seedNodeLeft
                             if isLeft else self._parameterNode.seedNodeRight)

            # Only reuse existing nodes when they belong to the current CT
            ct_prefix = volume_node.GetName()
            if existing_seg is not None and ct_prefix not in existing_seg.GetName():
                existing_seg  = None
                existing_seed = None

            # Posteriore Cutoff-Ebene
            cutoff_node = self._parameterNode.posteriorCutoffNode
            if cutoff_node is None or cutoff_node.GetNumberOfControlPoints() == 0:
                raise ValueError(
                    _("Bitte zuerst einen posterioren Cutoff-Punkt setzen "
                      "(Schaltfläche 'P' im Abschnitt 'Posteriore Cutoff-Ebene').")
                )
            _cutoff_pt = [0.0, 0.0, 0.0]
            cutoff_node.GetNthControlPointPositionWorld(0, _cutoff_pt)
            posterior_cutoff_ras = np.array(_cutoff_pt)

            # FM seed mode (only relevant for Fast Marching)
            seed_offset      = (None if self.ui.autoSeedCheckBox.isChecked()
                                else self.ui.seedOffsetSpinBox.value)
            seed_mode        = (1 if self.ui.rbSeedContralateral.isChecked()
                                else 2 if self.ui.rbSeedModelBased.isChecked() else 0)
            contra_seg_node  = None
            contra_seg_id    = None
            model_seed_node  = None
            model_seed_id    = None
            if method == 'fastmarching':
                if seed_mode == 1:
                    s        = "Left" if isLeft else "Right"
                    pos_node = getattr(self._parameterNode, f"contraPositionedNode{s}", None)
                    pos_tf   = getattr(self._parameterNode, f"contraPositionedTransform{s}", None)
                    if (pos_node is not None and slicer.mrmlScene.IsNodePresent(pos_node)
                            and pos_tf is not None and slicer.mrmlScene.IsNodePresent(pos_tf)):
                        model_seed_node = pos_node
                        pos_seg = pos_node.GetSegmentation()
                        if pos_seg.GetNumberOfSegments() > 0:
                            model_seed_id = pos_seg.GetNthSegmentID(0)
                    else:
                        contra_node = (self._parameterNode.segmentationNodeRight
                                       if isLeft else self._parameterNode.segmentationNodeLeft)
                        if contra_node is not None:
                            contra_seg_node = contra_node
                            contra_seg_id   = self._segIds.get(not isLeft)
                            if contra_seg_id is None:
                                seg = contra_node.GetSegmentation()
                                for i in range(seg.GetNumberOfSegments()):
                                    if seg.GetNthSegment(i).GetName() == "IntraorbitalVolume":
                                        contra_seg_id = seg.GetNthSegmentID(i)
                                        break
                if seed_mode == 2:
                    model_seed_node = (self._parameterNode.modelSeedNodeLeft if isLeft
                                       else self._parameterNode.modelSeedNodeRight)
                    if model_seed_node is not None:
                        seg = model_seed_node.GetSegmentation()
                        if seg.GetNumberOfSegments() > 0:
                            model_seed_id = seg.GetNthSegmentID(0)

            active_mrml   = self._parameterNode.parameterNode
            segment_color = self.logic._getColorForActiveParameterNode("Segmentation", active_mrml)

            result = self.logic.segmentIntraorbitalVolume(
                volume_node=volume_node,
                plane_model=plane_model,
                method=method,
                hu_min=self.ui.huMinSpinBox.value,
                hu_max=self.ui.huMaxSpinBox.value,
                radius_margin_mm=self.ui.radiusMarginSpinBox.value,
                seed_offset_mm=seed_offset,
                stopping_value=self.ui.stoppingValueSpinBox.value,
                speed_sigma=self.ui.speedSigmaSpinBox.value,
                posterior_boost=self.ui.posteriorBoostSpinBox.value,
                show_seed=self.ui.showSeedCheckBox.isChecked(),
                existing_segmentation_node=existing_seg,
                existing_seed_node=existing_seed,
                contralateral_seg_node=contra_seg_node,
                contralateral_seg_id=contra_seg_id,
                model_seed_node=model_seed_node,
                model_seed_id=model_seed_id,
                posterior_cutoff_ras=posterior_cutoff_ras,
                remove_satellites=self.ui.removeSatellitesCheckBox.isChecked(),
                min_satellite_diameter_mm=self.ui.satelliteDiamSpinBox.value,
                segment_color=segment_color,
                treat_air_as_soft_tissue=self.ui.treatAirAsSoftTissueCheckBox.isChecked(),
            )

            if isLeft:
                self._parameterNode.segmentationNodeLeft = result["segmentation_node"]
                self._parameterNode.segMethodLeft = method
                if result["seed_node"] is not None:
                    self._parameterNode.seedNodeLeft = result["seed_node"]
            else:
                self._parameterNode.segmentationNodeRight = result["segmentation_node"]
                self._parameterNode.segMethodRight = method
                if result["seed_node"] is not None:
                    self._parameterNode.seedNodeRight = result["seed_node"]

            self._placeInFolder(result["segmentation_node"], isLeft)
            if result["seed_node"] is not None:
                self._placeInFolder(result["seed_node"], isLeft)

            # Keep the manual selector in sync
            wasBlocked = self.ui.segNodeSelector.blockSignals(True)
            self.ui.segNodeSelector.setCurrentNode(result["segmentation_node"])
            self.ui.segNodeSelector.blockSignals(wasBlocked)

            # Measure final volume after postprocessing
            seg_node_final = result["segmentation_node"]
            seg_id_final   = result["segment_id"]
            final_vol_ml, final_vox = self._calculateVolumeFromSegment(seg_node_final, seg_id_final)

            self._segVolumes[isLeft] = final_vol_ml

            text = (
                _("<b>Intraorbital volume: {vol:.2f} ml</b><br>"
                  "Voxels: {vox}<br>"
                  "HU window: {hu_min} – {hu_max}").format(
                    vol=final_vol_ml,
                    vox=f"{final_vox:,}",
                    hu_min=self.ui.huMinSpinBox.value,
                    hu_max=self.ui.huMaxSpinBox.value,
                )
            )
            self._segTexts[isLeft] = text
            self.ui.segmentationResultLabel.setText(text)

            # Alle Slice-Ansichten auf das Zentrum der Segmentierung springen
            seg_node = result["segmentation_node"]
            bounds = [0.0] * 6
            seg_node.GetRASBounds(bounds)
            cx = (bounds[0] + bounds[1]) / 2
            cy = (bounds[2] + bounds[3]) / 2
            cz = (bounds[4] + bounds[5]) / 2
            for view_name in ["Red", "Green", "Yellow"]:
                sliceWidget = slicer.app.layoutManager().sliceWidget(view_name)
                if sliceWidget:
                    sliceWidget.sliceLogic().GetSliceNode().JumpSliceByCentering(cx, cy, cz)

            # 3D-Ansicht auf Segmentierungs-Zentrum verschieben (Kamera-Richtung beibehalten)
            threeDWidget = slicer.app.layoutManager().threeDWidget(0)
            if threeDWidget:
                threeDView = threeDWidget.threeDView()
                cameraNode = slicer.modules.cameras.logic().GetViewActiveCameraNode(
                    threeDView.mrmlViewNode()
                )
                if cameraNode:
                    fp = cameraNode.GetFocalPoint()
                    pos = cameraNode.GetPosition()
                    dx, dy, dz = cx - fp[0], cy - fp[1], cz - fp[2]
                    cameraNode.SetFocalPoint(cx, cy, cz)
                    cameraNode.SetPosition(pos[0]+dx, pos[1]+dy, pos[2]+dz)
                    cameraNode.ResetClippingRange()
                    threeDView.forceRender()

            # Interaction Handles des Modell-Seed-Transforms ausblenden
            tf_attr = "modelSeedTransformLeft" if isLeft else "modelSeedTransformRight"
            model_tf = getattr(self._parameterNode, tf_attr, None)
            if model_tf is not None and slicer.mrmlScene.IsNodePresent(model_tf):
                disp = model_tf.GetDisplayNode()
                if disp:
                    disp.SetEditorVisibility(False)

            # Interaction Handles des Contralateral-Positionierungs-Transforms ausblenden
            s = "Left" if isLeft else "Right"
            contra_pos_tf = getattr(self._parameterNode, f"contraPositionedTransform{s}", None)
            if contra_pos_tf is not None and slicer.mrmlScene.IsNodePresent(contra_pos_tf):
                disp = contra_pos_tf.GetDisplayNode()
                if disp:
                    disp.SetEditorVisibility(False)

            # Volume-Rendering ausblenden
            if self._vrDisplayNode is not None:
                self._vrDisplayNode.SetVisibility(False)
                self.ui.toggleVolumeRenderingButton.blockSignals(True)
                self.ui.toggleVolumeRenderingButton.setChecked(False)
                self.ui.toggleVolumeRenderingButton.setText(_("3D on"))
                self.ui.toggleVolumeRenderingButton.blockSignals(False)

            # Observer für manuelle Nachbearbeitung einrichten
            seg_node_new = result["segmentation_node"]
            seg_id_new   = result["segment_id"]
            old_node = self._segNodes[isLeft]
            old_obs  = self._segObservers[isLeft]

            if old_node is not None and old_obs is not None:
                old_node.RemoveObserver(old_obs)

            self._segNodes[isLeft] = seg_node_new
            self._segIds[isLeft]   = seg_id_new
            self._segObservers[isLeft] = seg_node_new.AddObserver(
                vtk.vtkCommand.ModifiedEvent,
                lambda c, e, side=isLeft: self._onSegmentModified(side),
            )

    def _nextStep(self) -> None:
        # if self.ui.stepsToolbox.currentIndex < self.ui.stepsToolbox.count - 1:
        #     self.ui.stepsToolbox.setCurrentIndex(self.ui.stepsToolbox.currentIndex + 1)
        return

    # ------------------------------------------------------------------
    # Signalhandler – Landmark workflow
    # ------------------------------------------------------------------

    def onLandmarkNodeSelectorChanged(self, node) -> None:
        if self._parameterNode is None:
            return
        isLeft = self._parameterNode.sideIsLeft
        # Remove observer from previous node
        old_entry = self._landmarkObservers.get(isLeft)
        if old_entry is not None:
            old_node, old_obs_list = old_entry
            for obs in old_obs_list:
                try:
                    old_node.RemoveObserver(obs)
                except Exception:
                    pass
            self._landmarkObservers[isLeft] = None
        # Store in PN
        if isLeft:
            self._parameterNode.rimLandmarkNodeLeft  = node
        else:
            self._parameterNode.rimLandmarkNodeRight = node
        # Attach new observers
        if node is not None:
            obs_defined = node.AddObserver(
                slicer.vtkMRMLMarkupsNode.PointPositionDefinedEvent,
                lambda c, e, side=isLeft: self._onLandmarkPointAdded(side),
            )
            obs_removed = node.AddObserver(
                slicer.vtkMRMLMarkupsNode.PointRemovedEvent,
                lambda c, e, side=isLeft: self._updateLandmarkUI(side),
            )
            self._landmarkObservers[isLeft] = (node, [obs_defined, obs_removed])
        self._updateLandmarkUI(isLeft, lm_node=node)

    def _makeLandmarkNode(self, isLeft: bool):
        """Creates and registers a fresh landmark node for the given side."""
        if self._parameterNode is None:
            return None
        pn_name = self._parameterNode.parameterNode.GetName()
        side_suffix = "L" if isLeft else "R"
        node = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLMarkupsFiducialNode",
            f"{pn_name}_OrbitalRimLandmarks_{side_suffix}",
        )
        node.SetMaximumNumberOfControlPoints(len(ORBITAL_RIM_LANDMARKS))
        node.CreateDefaultDisplayNodes()
        disp = node.GetDisplayNode()
        if disp:
            disp.SetGlyphScale(3.5)
            disp.SetTextScale(3.0)
            color = (0.2, 0.7, 1.0) if isLeft else (1.0, 0.55, 0.1)
            disp.SetSelectedColor(*color)
            disp.SetColor(*color)
        self._placeInFolder(node, isLeft)
        return node

    def onNewLandmarkNodeButton(self) -> None:
        if self._parameterNode is None:
            return
        isLeft = self._parameterNode.sideIsLeft
        node = self._makeLandmarkNode(isLeft)
        if node is None:
            return
        # Block the signal to prevent a double-trigger, then call the handler
        # explicitly — qMRMLNodeComboBox may have auto-selected the new scene node
        # before setCurrentNode runs, in which case currentNodeChanged never fires.
        wasBlocked = self.ui.landmarkNodeSelector.blockSignals(True)
        self.ui.landmarkNodeSelector.setCurrentNode(node)
        self.ui.landmarkNodeSelector.blockSignals(wasBlocked)
        self.onLandmarkNodeSelectorChanged(node)

    def onStartLandmarkButton(self) -> None:
        if self._parameterNode is None:
            return
        isLeft = self._parameterNode.sideIsLeft
        lm_node = (self._parameterNode.rimLandmarkNodeLeft if isLeft
                   else self._parameterNode.rimLandmarkNodeRight)
        if lm_node is None:
            node = self._makeLandmarkNode(isLeft)
            if node is None:
                return
            wasBlocked = self.ui.landmarkNodeSelector.blockSignals(True)
            self.ui.landmarkNodeSelector.setCurrentNode(node)
            self.ui.landmarkNodeSelector.blockSignals(wasBlocked)
            self.onLandmarkNodeSelectorChanged(node)
            lm_node = node
        n = lm_node.GetNumberOfControlPoints()
        if n >= len(ORBITAL_RIM_LANDMARKS):
            slicer.util.infoDisplay(_("All 10 landmarks are already placed."))
            return
        self._activateLandmarkPlacement(lm_node)

    def onLandmarkPrevButton(self) -> None:
        if self._parameterNode is None:
            return
        isLeft = self._parameterNode.sideIsLeft
        lm_node = (self._parameterNode.rimLandmarkNodeLeft if isLeft
                   else self._parameterNode.rimLandmarkNodeRight)
        if lm_node is None:
            return
        n_defined = lm_node.GetNumberOfDefinedControlPoints()
        if n_defined == 0:
            return
        # PointPositionDefinedEvent never fires on removal, so no observer dance needed.
        # Defined points occupy indices 0..n_defined-1 in sequential placement.
        lm_node.RemoveNthControlPoint(n_defined - 1)
        self._updateLandmarkUI(isLeft)

    def onCurveFromLandmarksButton(self) -> None:
        with slicer.util.tryWithErrorDisplay(_("Error creating curve from landmarks."), waitCursor=True):
            if self._parameterNode is None:
                return
            isLeft = self._parameterNode.sideIsLeft
            lm_node = (self._parameterNode.rimLandmarkNodeLeft if isLeft
                       else self._parameterNode.rimLandmarkNodeRight)
            if lm_node is None or lm_node.GetNumberOfControlPoints() < len(ORBITAL_RIM_LANDMARKS):
                slicer.util.warningDisplay(_("All 10 landmarks must be placed first."))
                return
            color = (0.2, 0.7, 1.0) if isLeft else (1.0, 0.55, 0.1)
            curve_node = self.logic.createCurveFromLandmarks(lm_node, color)
            self._curveNodes[isLeft] = curve_node
            self._placeInFolder(curve_node, isLeft)
            self.ui.curveSelector.blockSignals(True)
            self.ui.curveSelector.setCurrentNode(curve_node)
            self.ui.curveSelector.blockSignals(False)
            self.ui.createSurfaceButton.setEnabled(True)

    def _onLandmarkPointAdded(self, isLeft: bool) -> None:
        """Called after a landmark position is committed (PointPositionDefinedEvent).
        GetNumberOfDefinedControlPoints() is exact at this point (no off-by-one).
        Re-activates placement for the next point because SetNthControlPointLabel()
        can interrupt Slicer's persistent placement pipeline."""
        lm_node = (self._parameterNode.rimLandmarkNodeLeft if isLeft
                   else self._parameterNode.rimLandmarkNodeRight) if self._parameterNode else None
        if lm_node is None:
            return
        n = lm_node.GetNumberOfDefinedControlPoints()
        # Rename the just-confirmed point (index n-1) to its landmark ID
        if 1 <= n <= len(ORBITAL_RIM_LANDMARKS):
            lm_id = ORBITAL_RIM_LANDMARKS[n - 1][0]
            lm_node.SetNthControlPointLabel(n - 1, lm_id)
        self._updateLandmarkUI(isLeft)
        if n >= len(ORBITAL_RIM_LANDMARKS):
            # All done — exit placement
            interNode = slicer.app.applicationLogic().GetInteractionNode()
            interNode.SetPlaceModePersistence(0)
            interNode.SetCurrentInteractionMode(slicer.vtkMRMLInteractionNode.Select)
        else:
            # Re-activate so the next point can be placed immediately.
            # Safe because n < 10, so we are not past the last landmark.
            self._activateLandmarkPlacement(lm_node)

    def _activateLandmarkPlacement(self, lm_node) -> None:
        """Enters Slicer's *persistent* placement mode for the given fiducial node.
        PlaceModePersistence=1 keeps Slicer in Place mode after each confirmed click,
        so there is no need to re-activate between landmarks."""
        selNode = slicer.app.applicationLogic().GetSelectionNode()
        selNode.SetActivePlaceNodeID(lm_node.GetID())
        selNode.SetActivePlaceNodeClassName("vtkMRMLMarkupsFiducialNode")
        interNode = slicer.app.applicationLogic().GetInteractionNode()
        interNode.SetPlaceModePersistence(1)
        interNode.SetCurrentInteractionMode(slicer.vtkMRMLInteractionNode.Place)

    def _updateLandmarkUI(self, isLeft: bool, lm_node=None) -> None:
        """Refreshes landmark progress label, description, and button states.

        lm_node may be passed directly (e.g. when the PN has not been written yet);
        if omitted, the current node is read from the ParameterNode."""
        if self._parameterNode is None:
            return
        if lm_node is None:
            lm_node = (self._parameterNode.rimLandmarkNodeLeft if isLeft
                       else self._parameterNode.rimLandmarkNodeRight)
        # GetNumberOfDefinedControlPoints() excludes the live preview/ghost point
        # that follows the cursor in placement mode, unlike GetNumberOfControlPoints().
        n_placed = lm_node.GetNumberOfDefinedControlPoints() if lm_node is not None else 0
        total = len(ORBITAL_RIM_LANDMARKS)
        self.ui.landmarkProgressLabel.setText(f"{n_placed} / {total} placed")
        lines = []
        for i, (lid, lname, ldesc) in enumerate(ORBITAL_RIM_LANDMARKS):
            if i < n_placed:
                lines.append(f'<span style="color:#888;">{lid} – {lname}</span>')
            elif i == n_placed:
                lines.append(
                    f'<b>{lid} – {lname}</b>'
                    f'<br><i style="color:#555;">{ldesc}</i>'
                )
            else:
                lines.append(f'{lid} – {lname}')
        self.ui.landmarkCurrentLabel.setText("<br>".join(lines))
        has_node = lm_node is not None
        self.ui.landmarkPrevButton.setEnabled(has_node and n_placed > 0)
        self.ui.landmarkStartButton.setEnabled(has_node and n_placed < total)
        self.ui.curveFromLandmarksButton.setEnabled(has_node and n_placed >= total)

    def onClearSideButton(self) -> None:
        """Löscht alle Segmentierungs-Nodes der aktuell gewählten Seite nach Bestätigung."""
        if self._parameterNode is None:
            return
        isLeft = self._parameterNode.sideIsLeft
        side_name = "Links" if isLeft else "Rechts"
        s = "Left" if isLeft else "Right"
        side_suffix = "L" if isLeft else "R"

        answer = qt.QMessageBox.question(
            slicer.util.mainWindow(),
            _("Segmentierung löschen"),
            _("Alle Segmentierungs-Nodes für Seite {side} wirklich löschen?\n"
              "(Seed, Transforms, Segmentierung, Preview-Nodes)").format(side=side_name),
            qt.QMessageBox.Yes | qt.QMessageBox.No,
            qt.QMessageBox.No,
        )
        if answer != qt.QMessageBox.Yes:
            return

        def _remove(node):
            if node is not None and slicer.mrmlScene.IsNodePresent(node):
                slicer.mrmlScene.RemoveNode(node)

        # Segmentierung + abgeleitete Preview-Nodes (benannte Hilfsknoten vom FM)
        seg_node = getattr(self._parameterNode, f"segmentationNode{s}", None)
        if seg_node is not None:
            prefix = seg_node.GetName().replace("_IntraorbitalSeg", "") + "_"
            for preview_name in [
                f"{prefix}ContraMirror_Full",
                f"{prefix}ContraMirror_Shrunk",
                f"{prefix}ModelSeed_full",
                f"{prefix}ModelSeed",
            ]:
                _remove(slicer.mrmlScene.GetFirstNodeByName(preview_name))
        _remove(seg_node)
        setattr(self._parameterNode, f"segmentationNode{s}", None)

        # Seed-Fiducial
        _remove(getattr(self._parameterNode, f"seedNode{s}", None))
        setattr(self._parameterNode, f"seedNode{s}", None)

        # Modell-Seed + Transform-Kette (Centering + Haupt-Transform)
        model_seg = getattr(self._parameterNode, f"modelSeedNode{s}", None)
        if model_seg is not None and slicer.mrmlScene.IsNodePresent(model_seg):
            centering_id = model_seg.GetTransformNodeID()
            if centering_id:
                centering_node = slicer.mrmlScene.GetNodeByID(centering_id)
                model_tf = getattr(self._parameterNode, f"modelSeedTransform{s}", None)
                if centering_node and centering_node != model_tf:
                    _remove(centering_node)
        _remove(model_seg)
        setattr(self._parameterNode, f"modelSeedNode{s}", None)

        model_tf = getattr(self._parameterNode, f"modelSeedTransform{s}", None)
        _remove(model_tf)
        setattr(self._parameterNode, f"modelSeedTransform{s}", None)

        # Centering-Transform (Fallback, falls noch vorhanden)
        _remove(slicer.mrmlScene.GetFirstNodeByName(f"ModelSeedCentering_{side_suffix}"))

        # Positioniertes Gegenseiten-Segment + Transform-Kette
        contra_pos_node = getattr(self._parameterNode, f"contraPositionedNode{s}", None)
        if contra_pos_node is not None and slicer.mrmlScene.IsNodePresent(contra_pos_node):
            centering_id = contra_pos_node.GetTransformNodeID()
            if centering_id:
                centering_node = slicer.mrmlScene.GetNodeByID(centering_id)
                contra_pos_tf = getattr(self._parameterNode, f"contraPositionedTransform{s}", None)
                if centering_node and centering_node != contra_pos_tf:
                    _remove(centering_node)
        _remove(contra_pos_node)
        setattr(self._parameterNode, f"contraPositionedNode{s}", None)

        contra_pos_tf = getattr(self._parameterNode, f"contraPositionedTransform{s}", None)
        _remove(contra_pos_tf)
        setattr(self._parameterNode, f"contraPositionedTransform{s}", None)

        _remove(slicer.mrmlScene.GetFirstNodeByName(f"ContraPositionedCentering_{side_suffix}"))

        # Observer vom Segmentierungsknoten lösen
        seg_obs = self._segObservers.get(isLeft)
        if self._segNodes.get(isLeft) is not None and seg_obs is not None:
            try:
                self._segNodes[isLeft].RemoveObserver(seg_obs)
            except Exception:
                pass
        self._segObservers[isLeft] = None

        # Widget-Zustand zurücksetzen
        self._segTexts[isLeft] = ""
        self._segVolumes[isLeft] = None
        self._segNodes[isLeft] = None
        self._segIds[isLeft] = None

        self.ui.segmentationResultLabel.setText("")
        wasBlocked = self.ui.segNodeSelector.blockSignals(True)
        self.ui.segNodeSelector.setCurrentNode(None)
        self.ui.segNodeSelector.blockSignals(wasBlocked)

    def onOpenSegmentEditorButton(self) -> None:
        isLeft   = self._parameterNode.sideIsLeft if self._parameterNode else True
        seg_node = self._segNodes.get(isLeft)
        if seg_node is None:
            return
        volume_node = self.ui.ctVolumeSelector.currentNode()

        slicer.util.selectModule("SegmentEditor")
        qt.QApplication.processEvents()
        try:
            w = slicer.modules.segmenteditor.widgetRepresentation().self()
            w.editor.setSegmentationNode(seg_node)
            if volume_node:
                try:
                    w.editor.setSourceVolumeNode(volume_node)
                except AttributeError:
                    w.editor.setMasterVolumeNode(volume_node)
            w.editor.setActiveEffectByName("Scissors")
        except Exception as exc:
            logging.warning(f"SegmentEditor setup failed: {exc}")

    def onSegNodeSelectorChanged(self, node) -> None:
        if self._parameterNode is None:
            return
        isLeft = self._parameterNode.sideIsLeft
        s = "Left" if isLeft else "Right"
        setattr(self._parameterNode, f"segmentationNode{s}", node)
        if node is not None:
            seg = node.GetSegmentation()
            seg_id = None
            for i in range(seg.GetNumberOfSegments()):
                if seg.GetNthSegment(i).GetName() == "IntraorbitalVolume":
                    seg_id = seg.GetNthSegmentID(i)
                    break
            self._segNodes[isLeft] = node
            self._segIds[isLeft]   = seg_id
            # Calculate and display volume immediately
            if seg_id is not None:
                try:
                    vol_ml, voxel_count = self._calculateVolumeFromSegment(node, seg_id)
                    self._segVolumes[isLeft] = vol_ml
                    text = (
                        _("<b>Intraorbital volume: {vol:.2f} ml</b><br>"
                          "Voxels: {vox}").format(vol=vol_ml, vox=f"{voxel_count:,}")
                    )
                    self._segTexts[isLeft] = text
                    self.ui.segmentationResultLabel.setText(text)
                except Exception:
                    pass
            else:
                self._segTexts[isLeft] = ""
                self.ui.segmentationResultLabel.setText("")
        else:
            self._segNodes[isLeft] = None
            self._segIds[isLeft]   = None
            self._segTexts[isLeft] = ""
            self._segVolumes[isLeft] = None
            self.ui.segmentationResultLabel.setText("")

    def onCutoffNodeSelectorChanged(self, node) -> None:
        """Wird aufgerufen wenn der Nutzer einen anderen Cutoff-Node aus dem Dropdown wählt."""
        if self._parameterNode is None:
            return
        old_node = self._parameterNode.posteriorCutoffNode
        if old_node is not node and self._cutoffMarkupObserver is not None:
            try:
                old_node.RemoveObserver(self._cutoffMarkupObserver)
            except Exception:
                pass
            self._cutoffMarkupObserver = None
        self._parameterNode.posteriorCutoffNode = node
        if node is not None and self._cutoffMarkupObserver is None:
            self._cutoffMarkupObserver = node.AddObserver(
                vtk.vtkCommand.ModifiedEvent,
                lambda c, e: self._onCutoffMarkupModified(),
            )
        self._onCutoffMarkupModified()

    def onPlaceCutoffButton(self) -> None:
        """Erstellt einen neuen Cutoff-Punkt (oder leert den bestehenden) und aktiviert
        den Platzierungs-Modus.  Wird ein bestehender Node bereits genutzt, werden nur
        die Kontrollpunkte entfernt – der Node selbst bleibt erhalten."""
        if self._parameterNode is None:
            return
        existing = self._parameterNode.posteriorCutoffNode
        if existing is not None and slicer.mrmlScene.IsNodePresent(existing):
            node = existing
            node.RemoveAllControlPoints()
        else:
            node = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLMarkupsFiducialNode", "PosteriorCutoff"
            )
            disp = node.GetDisplayNode()
            disp.SetSelectedColor(1.0, 0.4, 0.0)
            disp.SetColor(1.0, 0.4, 0.0)
            disp.SetGlyphScale(3.0)
            # Selector setzen → löst onCutoffNodeSelectorChanged aus
            # → setzt parameterNode.posteriorCutoffNode + Observer
            self.ui.cutoffNodeSelector.setCurrentNode(node)
            self._placeUnderCTVolume(node)
        node.SetMaximumNumberOfControlPoints(1)
        selNode   = slicer.app.applicationLogic().GetSelectionNode()
        selNode.SetActivePlaceNodeID(node.GetID())
        selNode.SetActivePlaceNodeClassName("vtkMRMLMarkupsFiducialNode")
        interNode = slicer.app.applicationLogic().GetInteractionNode()
        interNode.SetCurrentInteractionMode(slicer.vtkMRMLInteractionNode.Place)

    def _syncSegIdAndTriggerUpdate(self, isLeft: bool, segmentationNode) -> None:
        """Aktualisiert _segIds/_segNodes und triggert die Volumen-Neuberechnung."""
        seg = segmentationNode.GetSegmentation()
        seg_id = seg.GetSegmentIdBySegmentName("IntraorbitalVolume")
        if seg_id:
            self._segIds[isLeft]  = seg_id
            self._segNodes[isLeft] = segmentationNode
        self._onSegmentModified(isLeft)

    def onRemoveExtrusionsButton(self, clicked: bool):
        isLeft = self._parameterNode.sideIsLeft
        segmentationNode = self._parameterNode.segmentationNodeLeft if isLeft else self._parameterNode.segmentationNodeRight
        segment_id = segmentationNode.GetSegmentation().GetSegmentIdBySegmentName("IntraorbitalVolume")

        self.logic.segmentationRemoveExtrusions(
            self._parameterNode.ctVolume,
            segmentationNode,
            segment_id
        )

        # After extrusion removal (morphological opening), automatically keep only
        # the correct island so any disconnected pieces are dropped.
        anchor = self._getIslandAnchor(isLeft)
        if anchor is not None:
            self.logic._keepIslandAtPoint(segmentationNode, segment_id, anchor)

        self._syncSegIdAndTriggerUpdate(isLeft, segmentationNode)

    def _getIslandAnchor(self, isLeft: bool):
        """Returns a RAS anchor point inside the correct orbital island, or None."""
        plane = (self._parameterNode.planeModelLeft if isLeft
                 else self._parameterNode.planeModelRight)
        volume_node = self.ui.ctVolumeSelector.currentNode()
        if plane is not None and volume_node is not None:
            centroid, normal = self.logic._getPlaneFromModel(plane)
            normal = self.logic._ensurePosteriorDirection(normal, centroid, volume_node)
            return centroid + 10.0 * normal
        seed_node = (self._parameterNode.seedNodeLeft if isLeft
                     else self._parameterNode.seedNodeRight)
        if seed_node is not None and seed_node.GetNumberOfControlPoints() > 0:
            pt = [0.0, 0.0, 0.0]
            seed_node.GetNthControlPointPositionWorld(0, pt)
            return np.array(pt)
        return None

    def onSelectIslandButton(self, clicked: bool):
        isLeft = self._parameterNode.sideIsLeft
        segmentationNode = (self._parameterNode.segmentationNodeLeft if isLeft
                            else self._parameterNode.segmentationNodeRight)
        if segmentationNode is None:
            return
        segment_id = segmentationNode.GetSegmentation().GetSegmentIdBySegmentName("IntraorbitalVolume")
        if not segment_id:
            return

        anchor = self._getIslandAnchor(isLeft)
        if anchor is None:
            slicer.util.warningDisplay(_("No orbital plane or seed available for island selection."))
            return

        self.logic._keepIslandAtPoint(segmentationNode, segment_id, anchor)
        self._syncSegIdAndTriggerUpdate(isLeft, segmentationNode)

    def onPerformCutoffButton(self, clicked: bool):
        isLeft = self._parameterNode.sideIsLeft
        segmentationNode = self._parameterNode.segmentationNodeLeft if isLeft else self._parameterNode.segmentationNodeRight
        segment_id = segmentationNode.GetSegmentation().GetSegmentIdBySegmentName("IntraorbitalVolume")
        mask_segment_id = segmentationNode.GetSegmentation().GetSegmentIdBySegmentName("Mask")

        self.logic.segmentationPerformCutoff(
            self._parameterNode.ctVolume,
            segmentationNode,
            segment_id,
            mask_segment_id
        )
        self._syncSegIdAndTriggerUpdate(isLeft, segmentationNode)

    def onRefreshVolumeButton(self, clicked: bool = False) -> None:
        if self._parameterNode is None:
            return
        isLeft = self._parameterNode.sideIsLeft
        seg_node = (self._parameterNode.segmentationNodeLeft if isLeft
                    else self._parameterNode.segmentationNodeRight)
        if seg_node is None:
            return
        self._syncSegIdAndTriggerUpdate(isLeft, seg_node)

    def _onCutoffMarkupModified(self) -> None:
        """Aktualisiert das Positions-Label der posterioren Cutoff-Ebene."""
        if self._parameterNode is None:
            return
        node = self._parameterNode.posteriorCutoffNode
        if node is None or node.GetNumberOfControlPoints() == 0:
            self.ui.cutoffPositionLabel.setText(_("nicht gesetzt"))
            return
        pt = [0.0, 0.0, 0.0]
        node.GetNthControlPointPositionWorld(0, pt)
        self.ui.cutoffPositionLabel.setText(
            f"R={pt[0]:.1f}  A={pt[1]:.1f}  S={pt[2]:.1f}"
        )

    def _onSegmentModified(self, isLeft: bool) -> None:
        self._volumeUpdateSide = isLeft
        self._volumeUpdateTimer.start()

    def _doVolumeUpdate(self) -> None:
        isLeft   = self._volumeUpdateSide
        if isLeft is None:
            return
        seg_node = self._segNodes.get(isLeft)
        seg_id   = self._segIds.get(isLeft)
        if seg_node is None or seg_id is None:
            return
        volume_ml, voxel_count = self._calculateVolumeFromSegment(seg_node, seg_id)
        self._segVolumes[isLeft] = volume_ml
        text = (
            _("<b>Intraorbital volume: {vol:.2f} ml</b>"
              " &nbsp;<i>(after manual editing)</i><br>"
              "Voxels: {vox}").format(vol=volume_ml, vox=f"{voxel_count:,}")
        )
        self._segTexts[isLeft] = text
        if self._parameterNode and self._parameterNode.sideIsLeft == isLeft:
            self.ui.segmentationResultLabel.setText(text)

    def _calculateVolumeFromSegment(self, seg_node, seg_id) -> tuple:
        try:
            seg     = seg_node.GetSegmentation()
            segment = seg.GetSegment(seg_id)
            if segment is None:
                return 0.0, 0
            binary_lm = segment.GetRepresentation("Binary labelmap")
            if binary_lm is None:
                return 0.0, 0
            import vtk.util.numpy_support as nps # pyright: ignore[reportMissingImports]
            scalars = binary_lm.GetPointData().GetScalars()
            if scalars is None:
                return 0.0, 0
            arr     = nps.vtk_to_numpy(scalars)
            spacing = binary_lm.GetSpacing()
            # After Segment Editor use, all segments share one collapsed labelmap
            # with distinct label values.  GetLabelValue() (Slicer ≥ 5.4) returns
            # the correct value; fall back to 1 for per-segment representations.
            try:
                label_value = segment.GetLabelValue()
            except AttributeError:
                label_value = 1
            voxel_count = int((arr == label_value).sum())
            volume_ml   = voxel_count * spacing[0] * spacing[1] * spacing[2] / 1000.0
            return volume_ml, voxel_count
        except Exception as exc:
            logging.warning(f"Volume recalculation failed: {exc}")
            return 0.0, 0

    # ------------------------------------------------------------------
    # Theme-aware styling
    # 
    # obsolete, stylesheets break theming in dark mode
    # ------------------------------------------------------------------

    def _refreshButtonStyles(self, *_) -> None:
        """Reapplies theme-sensitive stylesheets after a palette/style change."""
        defaultPalette = slicer.app.palette()

        themable_widgets = [
            self.ui.pageLandmarks,
            self.ui.pageVolumeSegmentation,
            self.ui.pageOrbitalSurface,
            self.ui.btnSideLeft,
            self.ui.btnSideRight
        ]

        # iterate through all widgets that need their colours adjusted
        for w in themable_widgets:
            if w.isChecked():
                # widget is selected/expanded, apply special palette
                activePalette = slicer.app.palette()
                activePalette.setColor(qt.QPalette.Button, qt.QColor("green"))
                activePalette.setColor(qt.QPalette.ButtonText, qt.QColor("white"))
                w.setPalette(activePalette)

                # make sure that the widgets children still have the default palette
                children = w.findChildren(qt.QWidget)
                for i in range(len(children)):
                    children[i].setPalette(defaultPalette)
            else:
                # widget not selected, apply default palette
                w.setPalette(defaultPalette)
           

    # ------------------------------------------------------------------
    # Subject Hierarchy helpers
    # ------------------------------------------------------------------

    def _shCTItemID(self) -> int:
        """SH item ID of the currently selected CT volume; creates the entry if missing.
        Falls back to scene root when no CT is selected."""
        shNode = slicer.vtkMRMLSubjectHierarchyNode.GetSubjectHierarchyNode(slicer.mrmlScene)
        ct_node = self.ui.ctVolumeSelector.currentNode()
        if ct_node is None:
            return shNode.GetSceneItemID()
        item_id = shNode.GetItemByDataNode(ct_node)
        if item_id == shNode.GetInvalidItemID():
            item_id = shNode.CreateItem(shNode.GetSceneItemID(), ct_node)
        return item_id

    def _getOrCreateSideFolder(self, isLeft: bool) -> int:
        shNode = slicer.vtkMRMLSubjectHierarchyNode.GetSubjectHierarchyNode(slicer.mrmlScene)
        parent_id = self._shCTItemID()
        pn_name = (self._parameterNode.parameterNode.GetName()
                   if self._parameterNode else "OrbitalVolumeWorkflow")
        side_label = "Links" if isLeft else "Rechts"
        folder_name = f"{pn_name} Orbital Volume Segmentation – {side_label}"
        child_ids = vtk.vtkIdList()
        shNode.GetItemChildren(parent_id, child_ids)
        for i in range(child_ids.GetNumberOfIds()):
            child_id = child_ids.GetId(i)
            if shNode.GetItemName(child_id) == folder_name:
                return child_id
        return shNode.CreateFolderItem(parent_id, folder_name)

    def _placeInFolder(self, node, isLeft: bool) -> None:
        if node is None:
            return
        shNode = slicer.vtkMRMLSubjectHierarchyNode.GetSubjectHierarchyNode(slicer.mrmlScene)
        folder_id = self._getOrCreateSideFolder(isLeft)
        item_id = shNode.GetItemByDataNode(node)
        if item_id == shNode.GetInvalidItemID():
            shNode.CreateItem(folder_id, node)
        else:
            shNode.SetItemParent(item_id, folder_id)

    def _placeUnderCTVolume(self, node) -> None:
        if node is None:
            return
        shNode = slicer.vtkMRMLSubjectHierarchyNode.GetSubjectHierarchyNode(slicer.mrmlScene)
        parent_id = self._shCTItemID()
        item_id = shNode.GetItemByDataNode(node)
        if item_id == shNode.GetInvalidItemID():
            shNode.CreateItem(parent_id, node)
        else:
            shNode.SetItemParent(item_id, parent_id)


# ═══════════════════════════════════════════════════════════════════════════════
# Logic
# ═══════════════════════════════════════════════════════════════════════════════

class OrbitalVolumeWorkflowModuleLogic(ScriptedLoadableModuleLogic):

    COLORS_SEGMENTATION = [
        (0.2, 0.8, 0.5), #green
        (0.5, 0.2, 0.8), #blue
        (0.8, 0.5, 0.2) #red
    ]

    COLORS_ORBITAL_PLANE = [
        (0.7, 1.0, 0.2), # green
        (0.2, 0.7, 1.0), # blue
        (1.0, 0.2, 0.7) # red
    ]

    COLOR_DEFAULT = (0.5, 0.5, 0.5) #grey

    def __init__(self) -> None:
        ScriptedLoadableModuleLogic.__init__(self)

    def getParameterNode(self):
        return OrbitalVolumeWorkflowModuleParameterNode(super().getParameterNode())

    def getAllParameterNodes(self):
        """Returns all vtkMRMLScriptedModuleNode instances belonging to this module."""
        result = []
        for i in range(slicer.mrmlScene.GetNumberOfNodesByClass("vtkMRMLScriptedModuleNode")):
            node = slicer.mrmlScene.GetNthNodeByClass(i, "vtkMRMLScriptedModuleNode")
            if node.GetAttribute("ModuleName") == "OrbitalVolumeWorkflowModule":
                result.append(node)
        return result

    # ------------------------------------------------------------------
    # Öffentliche API
    # ------------------------------------------------------------------

    def createOrbitalSurface(
        self,
        curve_node,
        method: str = "auto",
        subdivision_distance: float = 5.0,
        smooth_iterations: int = 0,
        color: tuple = None,
        opacity: float = 0.65,
    ):
        """
        Erstellt ein Surface-Mesh aus einer vtkMRMLMarkupsClosedCurveNode.

        Gibt den erzeugten vtkMRMLModelNode zurück.
        """
        print(f"\n{'='*55}")
        print(f"  Orbital Surface Generator")
        print(f"  Curve : {curve_node.GetName()}")

        pts = self._sampleCurvePoints(curve_node, subdivision_distance)
        print(f"  Points along curve   : {len(pts)}")

        if method == "auto":
            try:
                ci = self._concavityIndex(pts)
                print(f"  Concavity index      : {ci:.3f}")
                method = "delaunay" if ci > 1.08 else "fan"
                print(f"  Method (auto)        : {method}")
            except Exception:
                method = "delaunay"
                print("  Method (fallback)    : delaunay")
        else:
            print(f"  Method               : {method}")

        raw_poly = (
            self._delaunayTriangulation(pts)
            if method == "delaunay"
            else self._fanTriangulation(pts)
        )

        final_poly = self._postprocessSurface(raw_poly, smooth_iterations)

        model_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLModelNode")
        model_node.SetName(f"{curve_node.GetName()}_OrbitalPlane")
        model_node.SetAndObservePolyData(final_poly)
        model_node.CreateDefaultDisplayNodes()
    
        if color == None:
            color = self._getColorForActiveParameterNode("OrbitalPlane")

        disp = model_node.GetDisplayNode()
        disp.SetOpacity(opacity)
        disp.SetColor(*color)
        disp.SetBackfaceCulling(False)
        disp.SetRepresentation(2)
        disp.SetEdgeVisibility(True)

        print(f"  Triangles / Points   : {final_poly.GetNumberOfCells()} / {final_poly.GetNumberOfPoints()}")
        print(f"  Node name            : {model_node.GetName()}")
        print(f"{'='*55}\n")

        return model_node

    def createCurveFromLandmarks(self, landmark_node, color: tuple = (0.2, 0.7, 1.0)):
        """Creates a vtkMRMLMarkupsClosedCurveNode through all landmarks in L1→L10 order.

        The resulting curve can be passed directly to createOrbitalSurface() (Step 1).
        """
        n = landmark_node.GetNumberOfControlPoints()
        curve_node = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLMarkupsClosedCurveNode",
            landmark_node.GetName().replace("_OrbitalRimLandmarks_", "_OrbitalRim_"),
        )
        for i in range(n):
            pt = [0.0, 0.0, 0.0]
            landmark_node.GetNthControlPointPositionWorld(i, pt)
            curve_node.AddControlPoint(pt)
        curve_node.CreateDefaultDisplayNodes()
        disp = curve_node.GetDisplayNode()
        if disp:
            disp.SetColor(*color)
            disp.SetGlyphScale(3.0)
        return curve_node

    def segmentIntraorbitalVolume(
        self,
        volume_node,
        plane_model,
        method: str = 'threshold',
        hu_min: int = -200,
        hu_max: int = 300,
        radius_margin_mm: float = 5.0,
        # Fast Marching only
        seed_offset_mm=None,
        stopping_value: float = 25.0,
        speed_sigma: float = 70.0,
        posterior_boost: float = 1.5,
        show_seed: bool = False,
        existing_seed_node=None,
        contralateral_seg_node=None,
        contralateral_seg_id=None,
        model_seed_node=None,
        model_seed_id=None,
        # Common
        segment_color: tuple = None,
        existing_segmentation_node=None,
        remove_satellites: bool = True,
        min_satellite_diameter_mm: float = 3.0,
        posterior_cutoff_ras=None,
        treat_air_as_soft_tissue: bool = True,
    ) -> dict:
        """Segments the intraorbital volume by Fast Marching or HU threshold.

        method: 'fastmarching' | 'threshold'
        Returns a dict with keys: segmentation_node, segment_id, volume_ml,
        voxel_count, seed_node, voxel_exclusion_map.
        """
        label = "Fast Marching" if method == 'fastmarching' else "Threshold"
        print(f"\n{'='*60}")
        print(f"  Intraorbital Volume Segmentation ({label})")
        print(f"{'='*60}")
        print(f"  CT volume    : {volume_node.GetName()}")
        print(f"  Entry plane  : {plane_model.GetName()}")
        print(f"  HU window    : [{hu_min}, {hu_max}]")

        pd = qt.QProgressDialog(
            _("Vorbereitung..."), None, 0, 100, slicer.util.mainWindow()
        )
        pd.setWindowTitle(_("Intraorbitale Segmentierung"))
        pd.setWindowModality(qt.Qt.WindowModal)
        pd.setMinimumDuration(0)
        pd.show()
        pd.setValue(5)
        slicer.app.processEvents()

        centroid, normal = self._getPlaneFromModel(plane_model)
        normal = self._ensurePosteriorDirection(normal, centroid, volume_node)
        orbital_radius_mm = self._estimateOrbitalRadius(plane_model)

        if posterior_cutoff_ras is None:
            raise ValueError(_("Bitte zuerst einen posterioren Cutoff-Punkt setzen."))
        orbital_depth_mm = max(10.0, float(
            np.dot(np.array(posterior_cutoff_ras) - centroid, normal)
        ))

        print(f"  Centroid (RAS): {np.round(centroid, 1)}")
        print(f"  Normal vector : {np.round(normal, 3)}")
        print(f"  Orbital radius: {orbital_radius_mm:.1f} mm")
        print(f"  Orbital depth : {orbital_depth_mm:.1f} mm (from cutoff markup)")

        # --- Seed (Fast Marching only) ---
        seed_node = None
        seed_ras  = centroid  # fallback
        offset_used = 0.0
        hu_at_seed  = 0.0
        if method == 'fastmarching':
            ct_vol_name = volume_node.GetName()
            side_suffix = plane_model.GetName().replace("_OrbitalPlane", "")
            seed_name   = f"{ct_vol_name}_{side_suffix}_Seed"
            if seed_offset_mm is not None:
                seed_ras   = centroid + seed_offset_mm * normal
                hu_at_seed = self._getHUatRAS(seed_ras, volume_node) or 0.0
                offset_used = seed_offset_mm
            else:
                seed_ras, offset_used, hu_at_seed = self._findSeedPoint(
                    centroid, normal, volume_node, hu_min, hu_max
                )
            print(f"  Seed offset   : {offset_used:.1f} mm, HU={hu_at_seed:.0f}")
            seed_node = self._placeSeedFiducial(
                seed_ras, name=seed_name, existing_node=existing_seed_node
            )
            seed_node.GetDisplayNode().SetVisibility(1 if show_seed else 0)

        ct_vol_name   = volume_node.GetName()
        side_suffix   = plane_model.GetName().replace("_OrbitalPlane", "")
        seg_node_name = f"{ct_vol_name}_{side_suffix}_IntraorbitalSeg"

        if existing_segmentation_node is not None:
            segmentation_node = existing_segmentation_node
            segmentation_node.SetName(seg_node_name)
            segmentation_node.GetSegmentation().RemoveAllSegments()
            print(f"  Reusing segmentation node: {seg_node_name}")
        else:
            segmentation_node = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLSegmentationNode", seg_node_name
            )
            segmentation_node.CreateDefaultDisplayNodes()

        segmentation_node.SetReferenceImageGeometryParameterFromVolumeNode(volume_node)

        # --- Run the chosen algorithm ---
        if method == 'fastmarching':
            print(
                f"\n  Fast Marching "
                f"(stopping={stopping_value}, sigma={speed_sigma}, boost={posterior_boost}, "
                f"cylinder: {orbital_depth_mm} mm x r={orbital_radius_mm:.1f}+{radius_margin_mm} mm) ..."
            )
            pd.setLabelText(_("Fast Marching Segmentierung..."))
            pd.setValue(15)
            slicer.app.processEvents()

            seg_id, volume_ml, voxel_count, mask_segment_id = self._fastMarchingSegmentation(
                volume_node, segmentation_node, seed_ras,
                centroid, normal, orbital_radius_mm,
                orbital_depth_mm, radius_margin_mm,
                plane_model=plane_model,
                stopping_value=stopping_value,
                speed_sigma=speed_sigma,
                posterior_boost=posterior_boost,
                contralateral_seg_node=contralateral_seg_node,
                contralateral_seg_id=contralateral_seg_id,
                model_seed_node=model_seed_node,
                model_seed_id=model_seed_id,
                posterior_cutoff_ras=posterior_cutoff_ras,
                treat_air_as_soft_tissue=treat_air_as_soft_tissue,
            )
        else:
            print(
                f"\n  Threshold segmentation "
                f"(cylinder: {orbital_depth_mm} mm x r={orbital_radius_mm:.1f}+{radius_margin_mm} mm) ..."
            )
            pd.setLabelText(_("Schwellenwert-Segmentierung..."))
            pd.setValue(15)
            slicer.app.processEvents()

            seg_id, volume_ml, voxel_count, mask_segment_id = self._thresholdSegmentation(
                volume_node, segmentation_node,
                centroid, normal, orbital_radius_mm,
                orbital_depth_mm, radius_margin_mm,
                hu_min, hu_max,
                plane_model=plane_model,
                posterior_cutoff_ras=posterior_cutoff_ras,
                treat_air_as_soft_tissue=treat_air_as_soft_tissue,
            )

        pd.setLabelText(_("Segmentverwaltung..."))
        pd.setValue(55)
        slicer.app.processEvents()

        if segment_color is None:
            segment_color = self._getColorForActiveParameterNode("OrbitalPlane")

        seg = segmentation_node.GetSegmentation()
        segment_volume = seg.GetSegment(seg_id)
        segment_volume.SetName("IntraorbitalVolume")
        segment_volume.SetColor(*segment_color)

        mask_color = np.clip(np.multiply(segment_color, 2), 0, 1)
        segment_mask = seg.GetSegment(mask_segment_id)
        segment_mask.SetName("Mask")
        segment_mask.SetColor(mask_color)

        disp = segmentation_node.GetDisplayNode()
        disp.SetOpacity3D(0.5)
        disp.SetOpacity2DFill(0.4)
        disp.SetVisibility3D(True)
        segmentation_node.GetDisplayNode().SetSegmentVisibility(mask_segment_id, False)

        segmentation_node.CreateClosedSurfaceRepresentation()
        slicer.app.layoutManager().threeDWidget(0).threeDView().resetFocalPoint()

        print("\n  Postprocessing segmentation...")

        pd.setLabelText(_("Nachbearbeitung: Glättung..."))
        pd.setValue(62)
        slicer.app.processEvents()

        segmentEditorWidget = self._prepareSegmentEditor(volume_node, segmentation_node, seg_id)

        smoothing_factor = 0.8
        print(f"  - Smoothing (Joint Taubin, factor={smoothing_factor})...", end="")
        segmentEditorWidget.setActiveEffectByName("Smoothing")
        effect = segmentEditorWidget.activeEffect()
        effect.setParameter("SmoothingMethod", "JOINT_TAUBIN")
        effect.setParameter("JointTaubinSmoothingFactor", smoothing_factor)
        effect.parameterSetNode().SetMaskMode(slicer.vtkMRMLSegmentationNode.EditAllowedEverywhere)
        effect.parameterSetNode().SetOverwriteMode(slicer.vtkMRMLSegmentEditorNode.OverwriteNone)
        effect.self().onApply()
        time.sleep(1)
        print(" Done")

        pd.setLabelText(_("Nachbearbeitung: Closing..."))
        pd.setValue(74)
        slicer.app.processEvents()

        kernel_size = 4
        print(f"  - 1st. Smoothing (Closing, {kernel_size} mm)...", end="")
        self.segmentationPerformSmoothing(
            volume_node,
            segmentation_node,
            seg_id,
            "MORPHOLOGICAL_CLOSING",
            kernel_size,
            segmentEditorWidget
        )
        time.sleep(1)
        print(" Done")

        pd.setLabelText(_("Nachbearbeitung: HU-Schwellenwert..."))
        pd.setValue(80)
        slicer.app.processEvents()

        # Re-apply HU threshold inside the segment to remove bone introduced by smoothing
        # does not affect the soft tissue
        print(f"  - 1st. Re-threshold [-1024, {hu_max}] HU after smoothing...", end="")
        self.segmentationPerformThreshold(
            volume_node,
            segmentation_node,
            seg_id,
            -1024,
            hu_max,
            True,
            segmentEditorWidget
        )
        time.sleep(1)
        print(" Done")

        pd.setLabelText(_("Nachbearbeitung: Closing..."))
        pd.setValue(86)
        kernel_size = 2
        print(f"  - 2nd. Smoothing (Closing, {kernel_size} mm)...", end="")
        self.segmentationPerformSmoothing(
            volume_node,
            segmentation_node,
            seg_id,
            "MORPHOLOGICAL_CLOSING",
            kernel_size,
            segmentEditorWidget
        )
        time.sleep(1)
        print(" Done")

         # Re-apply HU threshold inside the segment to remove bone introduced by smoothing
        # does not affect the soft tissue
        pd.setLabelText(_("Nachbearbeitung: HU-Schwellenwert..."))
        pd.setValue(92)
        print(f"  - 2nd. Re-threshold [-1024, {hu_max+200}] HU after smoothing...", end="")
        self.segmentationPerformThreshold(
            volume_node,
            segmentation_node,
            seg_id,
            -1024,
            hu_max+100,
            True,
            segmentEditorWidget
        )
        time.sleep(1)
        print(" Done")

        pd.setValue(100)
        pd.close()
        slicer.app.processEvents()

        msgBox = qt.QMessageBox(qt.QMessageBox.Information,
                                _("Segmentation complete"),
                                _("Segmentation almost complete.<br /><br />"
                                  "Verify that the intraorbital space is fully filled. "
                                  "If not, adjust the parameters (HU min/max, Stopping Value) "
                                  "and re-run the segmentation.<br /><br />"
                                  "Once satisfied, click <b>Remove Extrusions</b> to remove "
                                  "protrusions and disconnected islands automatically.<br /><br />"
                                  "If isolated islands remain afterwards, click "
                                  "<b>Remove Islands</b> to keep only the correct one."))
        msgBox.addButton("OK", qt.QMessageBox.AcceptRole)
        msgBox.exec()

        print(f"\n{'─'*60}")
        print(f"  Threshold voxels       : {voxel_count:,}  ({volume_ml:.2f} ml before post-processing)")
        print(f"  Segment node           : {seg_node_name}")
        print(f"{'='*60}\n")

        slicer.util.setSliceViewerLayers(background=volume_node)

        return {
            "segmentation_node":   segmentation_node,
            "segment_id":          seg_id,
            "volume_ml":           volume_ml,
            "voxel_count":         voxel_count,
            "seed_node":           None,
            "voxel_exclusion_map": mask_segment_id,
        }

    def _prepareSegmentEditor(self, volume_node, segmentation_node, segment_id):
        segmentEditorWidget = slicer.modules.segmenteditor.widgetRepresentation().self().editor    
        segmentEditorWidget.setSegmentationNode(segmentation_node)
        segmentEditorWidget.setSourceVolumeNode(volume_node)
        segmentEditorNode = segmentEditorWidget.mrmlSegmentEditorNode()
        segmentEditorNode.SetSelectedSegmentID(segment_id)

        return segmentEditorWidget
    
    def segmentationRemoveExtrusions(self, volume_node, segmentation_node, segment_id, kernel_size = 8):
        # 4. Smoothing Opening 8-10 mm inside segment

        print(f"Smoothing (Opening) with a kernel size of {kernel_size} mm...", end="")

        segmentEditorWidget = self._prepareSegmentEditor(volume_node, segmentation_node, segment_id)
        segmentEditorWidget.setActiveEffectByName("Smoothing")
        effect = segmentEditorWidget.activeEffect()
        effect.setParameter("SmoothingMethod", "MORPHOLOGICAL_OPENING")
        effect.setParameter("KernelSizeMm", kernel_size)
        effect.parameterSetNode().SetMaskMode(slicer.vtkMRMLSegmentationNode.EditAllowedEverywhere)
        effect.self().onApply()
        print(" Done")

        segmentEditorWidget.setActiveEffectByName("Null")

    def segmentationKeepSelectedIsland(self, volume_node, segmentation_node, segment_id):
        segmentEditorWidget = self._prepareSegmentEditor(volume_node, segmentation_node, segment_id)

        segmentEditorWidget.setActiveEffectByName("Islands")
        effect = segmentEditorWidget.activeEffect()
        effect.setParameter("Operation", "KEEP_SELECTED_ISLAND")
        effect.parameterSetNode().SetMaskMode(slicer.vtkMRMLSegmentationNode.EditAllowedEverywhere)

    def segmentationPerformCutoff(self, volume_node, segmentation_node, segment_id, mask_segment_id):
        print(f"Performing anterior and posterior cutoff...", end="")
        
        segmentEditorWidget = self._prepareSegmentEditor(volume_node, segmentation_node, segment_id)

        segmentEditorWidget.setActiveEffectByName("Logical operators")
        effect = segmentEditorWidget.activeEffect()
        effect.setParameter("Operation","SUBTRACT")
        effect.setParameter("ModifierSegmentID", mask_segment_id)
        effect.parameterSetNode().SetMaskMode(slicer.vtkMRMLSegmentationNode.EditAllowedEverywhere)
        effect.self().onApply()

        print(" Done")

        # remove small irregularities after performing the cutoff
        self.segmentationRemoveExtrusions(volume_node, segmentation_node, segment_id, 2)

        segmentEditorWidget.setActiveEffectByName("Null")

    def segmentationPerformThreshold(self,
                                     volume_node,
                                     segmentation_node,
                                     segment_id,
                                     hu_min,
                                     hu_max,
                                     only_edit_inside = False,
                                     segmentEditorWidget = None):
        if segmentEditorWidget is None:
            segmentEditorWidget = self._prepareSegmentEditor(volume_node, segmentation_node, segment_id)
        
        segmentEditorWidget.setActiveEffectByName("Threshold")
        effect = segmentEditorWidget.activeEffect()
        effect.setParameter("MinimumThreshold", -1024)
        effect.setParameter("MaximumThreshold", hu_max)

        print(f"hu_max: {hu_max}")

        if only_edit_inside:
           effect.parameterSetNode().SetMaskMode(slicer.vtkMRMLSegmentationNode.EditAllowedInsideSingleSegment)
           effect.parameterSetNode().SetMaskSegmentID(segment_id)

        effect.parameterSetNode().SetMaskMode(slicer.vtkMRMLSegmentationNode.EditAllowedInsideVisibleSegments)
        effect.self().onApply()
        segmentEditorWidget.setActiveEffectByName("Null")

    def segmentationPerformSmoothing(self,
                                     volume_node,
                                     segmentation_node,
                                     segment_id,
                                     method,
                                     kernel_size,
                                     segmentEditorWidget = None):
        if segmentEditorWidget is None:
            segmentEditorWidget = self._prepareSegmentEditor(volume_node, segmentation_node, segment_id)
        
        segmentEditorWidget.setActiveEffectByName("Smoothing")
        effect = segmentEditorWidget.activeEffect()
        effect.setParameter("SmoothingMethod", method)
        effect.setParameter("KernelSizeMm", kernel_size)
        effect.parameterSetNode().SetMaskMode(slicer.vtkMRMLSegmentationNode.EditAllowedEverywhere)
        effect.parameterSetNode().SetOverwriteMode(slicer.vtkMRMLSegmentEditorNode.OverwriteNone)
        effect.self().onApply()
        segmentEditorWidget.setActiveEffectByName("Null")

    # ------------------------------------------------------------------
    # Hilfsfunktionen – Orbital Surface
    # ------------------------------------------------------------------

    def _sampleCurvePoints(self, curve_node, subdivision_distance: float):
        """Resamples the closed curve at uniform arc-length intervals and returns all control points as an (N,3) RAS array.
        Operates on a temporary copy so the original rim curve is never modified."""
        tmp = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLMarkupsClosedCurveNode", "_TmpSampleCurve"
        )
        try:
            tmp.Copy(curve_node)
            tmp.ResampleCurveWorld(subdivision_distance)
            n = tmp.GetNumberOfControlPoints()
            pts = np.zeros((n, 3))
            for i in range(n):
                p = [0.0, 0.0, 0.0]
                tmp.GetNthControlPointPositionWorld(i, p)
                pts[i] = p
        finally:
            slicer.mrmlScene.RemoveNode(tmp)
        return pts

    def _concavityIndex(self, pts):
        """Returns convex-hull-area / polygon-area; values > 1 indicate a concave (non-convex) curve, which selects fan over Delaunay triangulation."""
        from scipy.spatial import ConvexHull

        centroid = pts.mean(axis=0)
        centered = pts - centroid
        _, _, Vt = np.linalg.svd(centered, full_matrices=False)
        pts2d = centered @ Vt[:2].T

        x, y = pts2d[:, 0], pts2d[:, 1]
        poly_area = 0.5 * abs(np.dot(x, np.roll(y, -1)) - np.dot(y, np.roll(x, -1)))

        hull = ConvexHull(pts2d)
        hull_area = hull.volume  # in 2D ist .volume die Fläche

        return hull_area / poly_area if poly_area > 1e-6 else 1.0

    def _pcaProject(self, pts):
        """Projects 3D points onto their best-fit plane via SVD; returns 2D coordinates, centroid, and the rotation matrix Vt for back-projection."""
        centroid = pts.mean(axis=0)
        centered = pts - centroid
        _, _, Vt = np.linalg.svd(centered, full_matrices=False)
        return centered @ Vt[:2].T, centroid, Vt

    def _fanTriangulation(self, pts):
        """Connects every consecutive edge of the curve to the centroid, producing a star-shaped mesh; robust for concave or self-intersecting curves where Delaunay would produce artefacts."""
        n = len(pts)
        centroid = pts.mean(axis=0)

        vtk_pts = vtk.vtkPoints()
        for p in pts:
            vtk_pts.InsertNextPoint(p.tolist())
        vtk_pts.InsertNextPoint(centroid.tolist())
        center_idx = n

        triangles = vtk.vtkCellArray()
        for i in range(n):
            tri = vtk.vtkTriangle()
            tri.GetPointIds().SetId(0, center_idx)
            tri.GetPointIds().SetId(1, i)
            tri.GetPointIds().SetId(2, (i + 1) % n)
            triangles.InsertNextCell(tri)

        poly = vtk.vtkPolyData()
        poly.SetPoints(vtk_pts)
        poly.SetPolys(triangles)
        return poly

    def _delaunayTriangulation(self, pts):
        """Projects the curve onto its best-fit plane, runs constrained 2D Delaunay triangulation, then back-projects to 3D; produces a higher-quality mesh for convex curves than fan triangulation."""
        pts2d, centroid, Vt = self._pcaProject(pts)

        vtk_pts = vtk.vtkPoints()
        for p2 in pts2d:
            vtk_pts.InsertNextPoint(p2[0], p2[1], 0.0)

        n = len(pts2d)
        boundary = vtk.vtkCellArray()
        polyline = vtk.vtkPolyLine()
        polyline.GetPointIds().SetNumberOfIds(n + 1)
        for i in range(n):
            polyline.GetPointIds().SetId(i, i)
        polyline.GetPointIds().SetId(n, 0)
        boundary.InsertNextCell(polyline)

        boundary_poly = vtk.vtkPolyData()
        boundary_poly.SetPoints(vtk_pts)
        boundary_poly.SetLines(boundary)

        delaunay = vtk.vtkDelaunay2D()
        delaunay.SetInputData(boundary_poly)
        delaunay.SetSourceData(boundary_poly)
        delaunay.SetTolerance(0.001)
        delaunay.SetAlpha(0.0)
        delaunay.Update()

        result_2d = delaunay.GetOutput()

        result_pts = vtk.vtkPoints()
        for i in range(result_2d.GetNumberOfPoints()):
            p = result_2d.GetPoint(i)
            p3d = np.array([p[0], p[1]]) @ Vt[:2] + centroid
            result_pts.InsertNextPoint(p3d.tolist())

        out_poly = vtk.vtkPolyData()
        out_poly.SetPoints(result_pts)
        out_poly.SetPolys(result_2d.GetPolys())
        return out_poly

    def _postprocessSurface(self, poly_data, smooth_iterations: int = 0):
        """Optionally smooths the mesh with a windowed-sinc filter, then recomputes consistent point and cell normals required for correct rendering and ray-casting."""
        if smooth_iterations > 0:
            smoother = vtk.vtkWindowedSincPolyDataFilter()
            smoother.SetInputData(poly_data)
            smoother.SetNumberOfIterations(smooth_iterations)
            smoother.BoundarySmoothingOff()
            smoother.NonManifoldSmoothingOn()
            smoother.NormalizeCoordinatesOn()
            smoother.Update()
            poly_data = smoother.GetOutput()

        normals = vtk.vtkPolyDataNormals()
        normals.SetInputData(poly_data)
        normals.ComputePointNormalsOn()
        normals.ComputeCellNormalsOn()
        normals.AutoOrientNormalsOn()
        normals.ConsistencyOn()
        normals.Update()
        return normals.GetOutput()

    # ------------------------------------------------------------------
    # Hilfsfunktionen – Volumensegmentierung
    # ------------------------------------------------------------------

    def _estimateOrbitalRadius(self, plane_model) -> float:
        """Returns the maximum distance from the mesh centroid to any rim point; using max (not mean) ensures the prefilter sphere encloses the entire rim, including eccentric superior regions."""
        poly = plane_model.GetPolyData()
        n_pts = poly.GetNumberOfPoints()
        pts = np.array([poly.GetPoint(i) for i in range(n_pts)])
        centroid = pts.mean(axis=0)
        # max statt mean: Zylinder muss den gesamten Orbitarand umschließen,
        # da sonst exzentrische Randbereiche (z.B. superior) abgeschnitten werden
        return float(np.linalg.norm(pts - centroid, axis=1).max())

    def _getPlaneFromModel(self, model_node):
        """Fits a plane to all mesh vertices via SVD and returns (centroid, unit_normal); the least-significant singular vector (Vt[2]) is the best-fit normal."""
        poly = model_node.GetPolyData()
        pts = np.array([poly.GetPoint(i) for i in range(poly.GetNumberOfPoints())])
        centroid = pts.mean(axis=0)
        centered = pts - centroid
        _, _, Vt = np.linalg.svd(centered, full_matrices=False)
        return centroid, Vt[2]

    def _ensurePosteriorDirection(self, normal, centroid, volume_node):
        """Flips the plane normal if it points anteriorly (away from the volume centre); the Fast Marching depth axis must point into the orbit, not out of it."""
        bounds = [0.0] * 6
        volume_node.GetRASBounds(bounds)
        vol_center = np.array([
            (bounds[0] + bounds[1]) / 2,
            (bounds[2] + bounds[3]) / 2,
            (bounds[4] + bounds[5]) / 2,
        ])
        if np.dot(normal, vol_center - centroid) < 0:
            normal = -normal
        return normal

    def _rasToIjk(self, ras_point, volume_node):
        """Converts a RAS world coordinate to the nearest integer voxel index (IJK) using the volume's RAS-to-IJK matrix."""
        mat = vtk.vtkMatrix4x4()
        volume_node.GetRASToIJKMatrix(mat)
        ras_h = list(ras_point) + [1.0]
        ijk_h = mat.MultiplyPoint(ras_h)
        return np.round(ijk_h[:3]).astype(int)

    def _getHUatRAS(self, ras_point, volume_node):
        """Reads the scalar (HU) value from the CT volume at a given RAS position; returns None if the point lies outside the image extent."""
        ijk = self._rasToIjk(ras_point, volume_node)
        img = volume_node.GetImageData()
        dims = img.GetDimensions()
        if any(ijk[i] < 0 or ijk[i] >= dims[i] for i in range(3)):
            return None
        return img.GetScalarComponentAsFloat(int(ijk[0]), int(ijk[1]), int(ijk[2]), 0)

    def _findSeedPoint(self, centroid, normal, volume_node, hu_min, hu_max):
        """Walks along the orbital axis from the rim inward in 2 mm steps until a voxel within the soft-tissue HU window is found; raises if no candidate is found within 30 mm."""
        SEED_OFFSET_START_MM = 10.0
        SEED_OFFSET_STEP_MM  =  2.0
        SEED_OFFSET_MAX_MM   = 30.0

        offset = SEED_OFFSET_START_MM
        while offset <= SEED_OFFSET_MAX_MM:
            candidate = centroid + offset * normal
            hu = self._getHUatRAS(candidate, volume_node)
            if hu is not None and hu_min <= hu <= hu_max:
                return candidate, offset, hu
            offset += SEED_OFFSET_STEP_MM

        raise RuntimeError(
            _("No valid seed point found within the HU window [{hu_min}, {hu_max}] "
              "within {max_mm} mm. "
              "Adjust the HU window or seed offset.").format(
                hu_min=hu_min, hu_max=hu_max, max_mm=SEED_OFFSET_MAX_MM
            )
        )

    def _placeSeedFiducial(self, seed_ras, name: str = "OrbitalSeed", existing_node=None):
        """Creates a yellow fiducial markup at seed_ras, or repositions an existing one in-place to avoid cluttering the scene hierarchy on re-segmentation."""
        if existing_node is not None and slicer.mrmlScene.IsNodePresent(existing_node):
            existing_node.SetName(name)
            existing_node.RemoveAllControlPoints()
            existing_node.AddControlPoint(seed_ras.tolist())
            return existing_node
        node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLMarkupsFiducialNode")
        node.SetName(name)
        node.AddControlPoint(seed_ras.tolist())
        node.GetDisplayNode().SetSelectedColor(1.0, 0.8, 0.0)
        node.GetDisplayNode().SetGlyphScale(3.0)
        return node
    
    def _keepIslandAtPoint(self, segmentation_node, seg_id, anchor_ras):
        """Keep only the connected component of ``seg_id`` that contains
        ``anchor_ras`` (RAS).  Falls back to the largest component when the
        anchor is outside every island.  Short-circuits when ≤1 island exists.
        Modifies the binary labelmap in-place — segment ID does not change.
        """
        import SimpleITK as sitk
        import vtk.util.numpy_support as nps

        seg = segmentation_node.GetSegmentation()
        segment = seg.GetSegment(seg_id)
        if segment is None:
            return

        BL_NAME = "Binary labelmap"
        seg.CreateRepresentation(BL_NAME)
        binary_lm = segment.GetRepresentation(BL_NAME)
        if binary_lm is None:
            return
        scalars = binary_lm.GetPointData().GetScalars()
        if scalars is None:
            return

        try:
            label_value = segment.GetLabelValue()
        except AttributeError:
            label_value = 1

        dims = binary_lm.GetDimensions()          # (Nx, Ny, Nz)
        arr  = nps.vtk_to_numpy(scalars).reshape(dims[2], dims[1], dims[0])

        mask = (arr == label_value).astype(np.uint8)
        if mask.max() == 0:
            return

        cc_arr = sitk.GetArrayFromImage(
            sitk.ConnectedComponent(sitk.GetImageFromArray(mask))
        )
        n_islands = int(cc_arr.max())
        if n_islands <= 1:
            return  # nothing to select

        # GetImageToWorldMatrix maps VTK voxel index space → RAS world.
        # VTK indices start at the extent minimum, not at 0.
        # Invert to get RAS → VTK index, then subtract extent start to get numpy index.
        anchor_world = (float(anchor_ras[0]), float(anchor_ras[1]),
                        float(anchor_ras[2]), 1.0)

        i2w = vtk.vtkMatrix4x4()
        binary_lm.GetImageToWorldMatrix(i2w)
        w2i = vtk.vtkMatrix4x4()
        vtk.vtkMatrix4x4.Invert(i2w, w2i)
        idx = w2i.MultiplyPoint(anchor_world)

        ext = binary_lm.GetExtent()  # (ixMin, ixMax, iyMin, iyMax, izMin, izMax)
        ix = int(round(idx[0])) - ext[0]
        iy = int(round(idx[1])) - ext[2]
        iz = int(round(idx[2])) - ext[4]

        print(f"  Island anchor (RAS): {np.round(anchor_ras[:3], 1)}")
        print(f"  VTK index raw: ({idx[0]:.1f}, {idx[1]:.1f}, {idx[2]:.1f})  extent: {ext}")
        print(f"  Numpy index: ({ix}, {iy}, {iz})  array shape (z,y,x): {cc_arr.shape}")

        Nz, Ny, Nx = cc_arr.shape
        label_at_anchor = 0
        if 0 <= ix < Nx and 0 <= iy < Ny and 0 <= iz < Nz:
            label_at_anchor = int(cc_arr[iz, iy, ix])

        if label_at_anchor == 0:
            uniq, counts = np.unique(cc_arr[cc_arr > 0], return_counts=True)
            label_at_anchor = int(uniq[np.argmax(counts)])
            print(f"  Island auto-select: anchor outside all islands "
                  f"— keeping largest of {n_islands}")
        else:
            print(f"  Island auto-select: island {label_at_anchor} of {n_islands}")

        result = np.where(cc_arr == label_at_anchor, label_value, 0).astype(arr.dtype)
        new_sc = nps.numpy_to_vtk(result.flatten(order='C'), deep=True)
        new_sc.SetName(scalars.GetName())
        binary_lm.GetPointData().SetScalars(new_sc)
        binary_lm.Modified()
        segmentation_node.Modified()
        segmentation_node.CreateClosedSurfaceRepresentation()


    def _fastMarchingSegmentation(
        self,
        volume_node, segmentation_node, seed_ras,
        centroid, normal, orbital_radius_mm,
        orbital_depth_mm, radius_margin_mm,
        plane_model=None,
        stopping_value: float = 25.0,
        speed_sigma: float = 70.0,
        posterior_boost: float = 1.5,
        contralateral_seg_node=None,
        contralateral_seg_id=None,
        model_seed_node=None,
        model_seed_id=None,
        posterior_cutoff_ras=None,
        treat_air_as_soft_tissue: bool = True,
    ):
        """Segments the intraorbital volume via Fast Marching on a HU-derived speed image."""
        import SimpleITK as sitk
        import sitkUtils
        from scipy.interpolate import griddata

        c_lps = np.array([-centroid[0], -centroid[1],  centroid[2]])
        n_lps = np.array([-normal[0],   -normal[1],    normal[2]])
        max_extent = orbital_radius_mm + radius_margin_mm + orbital_depth_mm

        sitk_image_full = sitkUtils.PullVolumeFromSlicer(volume_node)
        sitk_image = self._cropSitkToOrbitalRegion(sitk_image_full, c_lps, max_extent + 10.0)
        del sitk_image_full

        hu_arr = sitk.GetArrayFromImage(sitk.Cast(sitk_image, sitk.sitkFloat32))
        print(f"  Crop size: {sitk_image.GetSize()} voxels "
              f"(spacing {np.round(sitk_image.GetSpacing(), 2)})")

        speed_arr = np.exp(-(hu_arr / speed_sigma) ** 2).astype(np.float32)

        origin    = np.array(sitk_image.GetOrigin())
        spacing   = np.array(sitk_image.GetSpacing())
        direction = np.array(sitk_image.GetDirection()).reshape(3, 3)

        Nz, Ny, Nx = hu_arr.shape
        IZ, IY, IX = np.meshgrid(np.arange(Nz), np.arange(Ny), np.arange(Nx), indexing='ij')
        idx_flat = np.stack([IX.flatten(), IY.flatten(), IZ.flatten()], axis=1).astype(np.float64)
        lps_pts = origin + (direction @ (idx_flat * spacing).T).T

        rel     = lps_pts - c_lps
        depth   = rel @ n_lps
        lateral = np.linalg.norm(rel - np.outer(depth, n_lps), axis=1)

        if plane_model is not None:
            world_up = np.array([0.0, 0.0, 1.0])
            u_vec = world_up - np.dot(world_up, n_lps) * n_lps
            if np.linalg.norm(u_vec) < 1e-6:
                u_vec = np.array([1.0, 0.0, 0.0]) - np.dot(np.array([1.0, 0.0, 0.0]), n_lps) * n_lps
            u_vec /= np.linalg.norm(u_vec)
            v_vec = np.cross(n_lps, u_vec)

            poly = plane_model.GetPolyData()
            mesh_ras = np.array([poly.GetPoint(i) for i in range(poly.GetNumberOfPoints())])
            mesh_lps = mesh_ras * np.array([-1.0, -1.0, 1.0])
            mesh_rel = mesh_lps - c_lps
            mesh_d   = mesh_rel @ n_lps
            mesh_u   = mesh_rel @ u_vec
            mesh_v   = mesh_rel @ v_vec

            centroid_dist = np.linalg.norm(rel, axis=1)
            in_region = (
                (centroid_dist <= max_extent) &
                (depth >= -5.0) &
                (depth <= orbital_depth_mm + 5.0)
            )

            rel_in       = rel[in_region]
            depth_in_reg = depth[in_region]
            vu = rel_in @ u_vec
            vv = rel_in @ v_vec

            mesh_depth_at_voxel = griddata(
                np.column_stack([mesh_u, mesh_v]),
                mesh_d,
                np.column_stack([vu, vv]),
                method='linear', fill_value=np.nan,
            )
            nan_mask = np.isnan(mesh_depth_at_voxel)
            if nan_mask.any():
                mesh_depth_at_voxel[nan_mask] = griddata(
                    np.column_stack([mesh_u, mesh_v]),
                    mesh_d,
                    np.column_stack([vu[nan_mask], vv[nan_mask]]),
                    method='nearest',
                )

            speed_barrier = (lateral > orbital_radius_mm + radius_margin_mm).reshape(Nz, Ny, Nx)
            cutoff_mask = ~in_region
            cutoff_mask[in_region] = (
                (depth_in_reg < mesh_depth_at_voxel) |
                (depth_in_reg > orbital_depth_mm)
            )
            cutoff_mask = cutoff_mask.reshape(Nz, Ny, Nx)
        else:
            speed_barrier = (lateral > orbital_radius_mm + radius_margin_mm).reshape(Nz, Ny, Nx)
            cutoff_mask = (
                (depth <= 0) |
                (depth > orbital_depth_mm) |
                (lateral > orbital_radius_mm + radius_margin_mm)
            ).reshape(Nz, Ny, Nx)

        if posterior_cutoff_ras is not None:
            cutoff_lps_pt = np.array([
                -posterior_cutoff_ras[0],
                -posterior_cutoff_ras[1],
                 posterior_cutoff_ras[2],
            ])
            depth_from_cutoff = (lps_pts - cutoff_lps_pt) @ n_lps
            cutoff_mask = cutoff_mask | (depth_from_cutoff.reshape(Nz, Ny, Nx) > 0)

        if treat_air_as_soft_tissue:
            orbital_air = (~speed_barrier) & (hu_arr < -400)
            speed_arr[orbital_air] = 1.0

        depth_norm = np.clip(depth / orbital_depth_mm, 0.0, 1.0)
        posterior_factor = (1.0 + posterior_boost * depth_norm).reshape(Nz, Ny, Nx)
        speed_arr = np.clip(speed_arr * posterior_factor, 1e-4, None)
        speed_arr[speed_barrier] = 1e-4

        speed_sitk = sitk.GetImageFromArray(speed_arr)
        speed_sitk.CopyInformation(sitk_image)

        s_lps = [-float(seed_ras[0]), -float(seed_ras[1]), float(seed_ras[2])]
        seed_idx = speed_sitk.TransformPhysicalPointToIndex(s_lps)
        print(f"  FM seed index (LPS): {seed_idx}")

        name_prefix = segmentation_node.GetName().replace("_IntraorbitalSeg", "") + "_"
        if contralateral_seg_node is not None and contralateral_seg_id is not None:
            region_points = self._prepareContralateralSeed(
                contralateral_seg_node, contralateral_seg_id, volume_node, name_prefix
            )
        elif model_seed_node is not None and model_seed_id is not None:
            region_points = self._prepareModelSeed(
                model_seed_node, model_seed_id, volume_node, name_prefix
            )
        else:
            region_points = []

        if posterior_cutoff_ras is not None and region_points:
            cutoff_lps_pre = np.array([
                -posterior_cutoff_ras[0],
                -posterior_cutoff_ras[1],
                 posterior_cutoff_ras[2],
            ])
            pts_ijk = np.array(region_points, dtype=np.float64)
            pts_lps = origin + (direction @ (pts_ijk * spacing).T).T
            keep = (pts_lps - cutoff_lps_pre) @ n_lps <= 0
            n_before = len(region_points)
            region_points = [p for p, k in zip(region_points, keep) if k]
            print(f"  Cutoff filter trial points: {len(region_points)}/{n_before} kept")

        trial_points = region_points if region_points else [seed_idx]
        if seed_idx not in trial_points:
            trial_points = [seed_idx] + trial_points

        fm = sitk.FastMarchingImageFilter()
        fm.SetStoppingValue(stopping_value)
        fm.SetTrialPoints(trial_points)
        arrival = fm.Execute(speed_sitk)

        arrival_arr = sitk.GetArrayFromImage(arrival)
        binary_arr  = (arrival_arr < stopping_value).astype(np.uint8)

        voxel_count = int(binary_arr.sum())
        sp = sitk_image.GetSpacing()
        volume_ml = voxel_count * sp[0] * sp[1] * sp[2] / 1000.0
        print(f"  Fast Marching: {voxel_count:,} voxels, {volume_ml:.2f} ml")

        seg_id          = self._convertArrayToSegment(binary_arr, sitk_image, segmentation_node)
        mask_segment_id = self._convertArrayToSegment(cutoff_mask.astype(np.uint8), sitk_image, segmentation_node)

        return seg_id, volume_ml, voxel_count, mask_segment_id

    def _thresholdSegmentation(
        self,
        volume_node, segmentation_node,
        centroid, normal, orbital_radius_mm,
        orbital_depth_mm, radius_margin_mm,
        hu_min: float, hu_max: float,
        plane_model=None,
        posterior_cutoff_ras=None,
        treat_air_as_soft_tissue: bool = True,
    ):
        """Segments the intraorbital volume by HU threshold within a geometric mask.

        The mask consists of a lateral cylinder (radius = orbital_radius_mm +
        radius_margin_mm), the anterior boundary defined by the entry-plane mesh,
        the posterior depth limit (orbital_depth_mm), and an optional posterior
        cutoff point.  Every voxel inside the mask whose HU is in [hu_min, hu_max]
        is included in the segment.
        """
        import SimpleITK as sitk
        import sitkUtils
        from scipy.interpolate import griddata

        # --- 1. Crop CT to orbital bounding box ---
        c_lps = np.array([-centroid[0], -centroid[1],  centroid[2]])
        n_lps = np.array([-normal[0],   -normal[1],    normal[2]])
        max_extent = orbital_radius_mm + radius_margin_mm + orbital_depth_mm

        # If a posterior cutoff point is given, extend the crop so it reaches
        # 10 mm beyond that point (its 3D distance from the centroid is the
        # relevant metric because _cropSitkToOrbitalRegion uses a cubic box).
        crop_radius = max_extent + 0.0
        if posterior_cutoff_ras is not None:
            cutoff_lps = np.array([
                -posterior_cutoff_ras[0],
                -posterior_cutoff_ras[1],
                 posterior_cutoff_ras[2],
            ])
            cutoff_dist = float(np.linalg.norm(cutoff_lps - c_lps))
            crop_radius = max(crop_radius, cutoff_dist + 0.0)

        sitk_image_full = sitkUtils.PullVolumeFromSlicer(volume_node)
        sitk_image = self._cropSitkToOrbitalRegion(sitk_image_full, c_lps, crop_radius)
        del sitk_image_full

        hu_arr = sitk.GetArrayFromImage(sitk.Cast(sitk_image, sitk.sitkFloat32))
        print(f"  Crop size: {sitk_image.GetSize()} voxels "
              f"(spacing {np.round(sitk_image.GetSpacing(), 2)})")

        # --- 2. Physical LPS coordinates for every voxel ---
        origin    = np.array(sitk_image.GetOrigin())
        spacing   = np.array(sitk_image.GetSpacing())
        direction = np.array(sitk_image.GetDirection()).reshape(3, 3)

        Nz, Ny, Nx = hu_arr.shape
        IZ, IY, IX = np.meshgrid(np.arange(Nz), np.arange(Ny), np.arange(Nx), indexing='ij')
        idx_flat = np.stack([IX.flatten(), IY.flatten(), IZ.flatten()], axis=1).astype(np.float64)
        lps_pts = origin + (direction @ (idx_flat * spacing).T).T

        # --- 3. Orbital coordinates: axial depth and lateral distance ---
        rel     = lps_pts - c_lps
        depth   = rel @ n_lps
        lateral = np.linalg.norm(rel - np.outer(depth, n_lps), axis=1)

        # --- 4. Lateral cylinder — applied during thresholding ---
        outside_cylinder = (lateral > orbital_radius_mm + radius_margin_mm).reshape(Nz, Ny, Nx)

        # --- 5. Cutoff mask (entry plane + depth limit + posterior cutoff point) ---
        # Computed here but NOT applied to the threshold — exported as a "Mask"
        # segment and subtracted via "Perform Cutoff" / "Finish Segmentation".
        if plane_model is not None:
            world_up = np.array([0.0, 0.0, 1.0])
            u_vec = world_up - np.dot(world_up, n_lps) * n_lps
            if np.linalg.norm(u_vec) < 1e-6:
                u_vec = np.array([1.0, 0.0, 0.0]) - np.dot(np.array([1.0, 0.0, 0.0]), n_lps) * n_lps
            u_vec /= np.linalg.norm(u_vec)
            v_vec = np.cross(n_lps, u_vec)

            poly = plane_model.GetPolyData()
            mesh_ras = np.array([poly.GetPoint(i) for i in range(poly.GetNumberOfPoints())])
            mesh_lps = mesh_ras * np.array([-1.0, -1.0, 1.0])
            mesh_rel = mesh_lps - c_lps
            mesh_d   = mesh_rel @ n_lps
            mesh_u   = mesh_rel @ u_vec
            mesh_v   = mesh_rel @ v_vec

            centroid_dist = np.linalg.norm(rel, axis=1)
            in_region = (
                (centroid_dist <= max_extent) &
                (depth >= -5.0) &
                (depth <= orbital_depth_mm + 5.0)
            )

            rel_in       = rel[in_region]
            depth_in_reg = depth[in_region]
            vu = rel_in @ u_vec
            vv = rel_in @ v_vec

            mesh_depth_at_voxel = griddata(
                np.column_stack([mesh_u, mesh_v]),
                mesh_d,
                np.column_stack([vu, vv]),
                method='linear', fill_value=np.nan,
            )
            nan_mask = np.isnan(mesh_depth_at_voxel)
            if nan_mask.any():
                mesh_depth_at_voxel[nan_mask] = griddata(
                    np.column_stack([mesh_u, mesh_v]),
                    mesh_d,
                    np.column_stack([vu[nan_mask], vv[nan_mask]]),
                    method='nearest',
                )

            cutoff_mask = ~in_region
            cutoff_mask[in_region] = (
                (depth_in_reg < mesh_depth_at_voxel) |  # anterior to entry plane
                (depth_in_reg > orbital_depth_mm)        # beyond depth limit
            )
            cutoff_mask = cutoff_mask.reshape(Nz, Ny, Nx)
        else:
            cutoff_mask = (
                (depth <= 0) | (depth > orbital_depth_mm)
            ).reshape(Nz, Ny, Nx)

        # Posterior cutoff point: plane perpendicular to orbital axis through markup
        if posterior_cutoff_ras is not None:
            cutoff_lps_pt = np.array([
                -posterior_cutoff_ras[0],
                -posterior_cutoff_ras[1],
                 posterior_cutoff_ras[2],
            ])
            depth_from_cutoff = (lps_pts - cutoff_lps_pt) @ n_lps
            cutoff_mask = cutoff_mask | (depth_from_cutoff.reshape(Nz, Ny, Nx) > 0)

        # --- 6. Apply HU threshold within the lateral cylinder only ---
        hu_in_range = (hu_arr >= hu_min) & (hu_arr <= hu_max)
        if treat_air_as_soft_tissue:
            hu_in_range = hu_in_range | (hu_arr < -400)
        binary_arr = ((~outside_cylinder) & hu_in_range).astype(np.uint8)

        voxel_count = int(binary_arr.sum())
        sp = sitk_image.GetSpacing()
        volume_ml = voxel_count * sp[0] * sp[1] * sp[2] / 1000.0
        print(f"  Threshold [{hu_min}, {hu_max}] HU: {voxel_count:,} voxels, {volume_ml:.2f} ml")

        # --- 7. Import results as segments ---
        seg_id         = self._convertArrayToSegment(binary_arr, sitk_image, segmentation_node)
        mask_segment_id = self._convertArrayToSegment(cutoff_mask.astype(np.uint8), sitk_image, segmentation_node)

        return seg_id, volume_ml, voxel_count, mask_segment_id

    def _cropSitkToOrbitalRegion(self, sitk_image, c_lps, radius_mm):
        """Crops a SimpleITK image to a cube of ±radius_mm around c_lps (LPS).

        Preserves Origin/Spacing/Direction so all subsequent physical-coordinate
        calculations remain correct.  Returns the original image unchanged if
        the crop cannot be computed (e.g. centroid outside FOV).
        """
        import SimpleITK as sitk
        import itertools

        img_size = sitk_image.GetSize()  # (Nx, Ny, Nz) in SimpleITK x,y,z order

        # Build all 8 corners of the physical bounding box and map to voxel indices.
        # Using all corners handles oblique image orientations correctly.
        corners = list(itertools.product(
            [c_lps[0] - radius_mm, c_lps[0] + radius_mm],
            [c_lps[1] - radius_mm, c_lps[1] + radius_mm],
            [c_lps[2] - radius_mm, c_lps[2] + radius_mm],
        ))
        indices = [sitk_image.TransformPhysicalPointToIndex(c) for c in corners]

        crop_start = [max(0, min(idx[d] for idx in indices)) for d in range(3)]
        crop_end   = [min(img_size[d], max(idx[d] for idx in indices) + 1) for d in range(3)]
        crop_size  = [crop_end[d] - crop_start[d] for d in range(3)]

        if any(s <= 0 for s in crop_size):
            return sitk_image

        return sitk.RegionOfInterest(sitk_image, size=crop_size, index=crop_start)

    def _convertArrayToSegment(self, binary_arr, sitk_image, segmentation_node):
        """Converts a 3-dimensional array to a 3D Slicer Segment

        :param binary_arr: 3-dimensional array
        :param sitk_image: ITK-Image object, that contains the volume properties
        :param segmentation_node: 3D Slicer segmentation node, that the segment will be added to
        :return: Segment-ID of the newly created Segment
        """        
        binary_sitk = sitk.GetImageFromArray(binary_arr)
        binary_sitk.CopyInformation(sitk_image)

        # create Segment for intraorbital Volume
        lm_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLabelMapVolumeNode", "TmpFMLabelmap")
        sitkUtils.PushVolumeToSlicer(binary_sitk, lm_node)
        slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
            lm_node, segmentation_node
        )
        slicer.mrmlScene.RemoveNode(lm_node)

        seg = segmentation_node.GetSegmentation()
        seg_ids = vtk.vtkStringArray()
        seg.GetSegmentIDs(seg_ids)
        seg_id = seg_ids.GetValue(seg_ids.GetNumberOfValues() - 1)

        return seg_id

    def _smoothSegment(self, segmentation_node, seg_id, kernel_size_mm: float = 3.0, method: str = "closing"):
        """Applies morphological closing or opening to the binary labelmap using an ellipsoidal kernel scaled to physical mm.

        The array is zero-padded before the operation so that scipy's border
        treatment (outside = background) does not erode voxels at the tight
        bounding-box edge, which would create an artificial flat cut.
        """
        import vtk.util.numpy_support as nps # pyright: ignore[reportMissingImports]
        from scipy import ndimage

        seg = segmentation_node.GetSegmentation()
        binary_lm = seg.GetSegment(seg_id).GetRepresentation("Binary labelmap")
        if binary_lm is None:
            raise RuntimeError(_("Segment has no binary labelmap representation."))

        dims    = binary_lm.GetDimensions()
        spacing = binary_lm.GetSpacing()
        scalars = binary_lm.GetPointData().GetScalars()

        flat      = nps.vtk_to_numpy(scalars)
        vol_array = flat.reshape(dims[2], dims[1], dims[0])

        r  = kernel_size_mm / 2.0
        rz = max(1, int(round(r / spacing[2])))
        ry = max(1, int(round(r / spacing[1])))
        rx = max(1, int(round(r / spacing[0])))
        z, y, x = np.ogrid[-rz:rz+1, -ry:ry+1, -rx:rx+1]
        struct = (z/rz)**2 + (y/ry)**2 + (x/rx)**2 <= 1.0

        binary = vol_array > 0
        # Padding verhindert Boundary-Artefakte: scipy behandelt außerhalb des
        # Arrays als Hintergrund (0), was die Erosionsphase alle Randvoxel
        # entfernt und einen künstlichen flachen Schnitt am Array-Rand erzeugt.
        pad = max(rz, ry, rx) + 1
        padded = np.pad(binary, pad, mode='constant', constant_values=0)
        if method == "closing":
            result_padded = ndimage.binary_closing(padded, structure=struct)
        elif method == "opening":
            result_padded = ndimage.binary_opening(padded, structure=struct)
        else:
            raise ValueError(_("Unknown method: {method}").format(method=repr(method)))
        result = result_padded[pad:-pad, pad:-pad, pad:-pad]

        flat_out = result.astype(np.uint8).flatten().astype(flat.dtype)
        scalars.DeepCopy(nps.numpy_to_vtk(flat_out))
        binary_lm.Modified()
        segmentation_node.Modified()

    def _maskSegmentByHU(self, segmentation_node, seg_id, volume_node, hu_max: float) -> None:
        """Removes all voxels from the segment whose CT value exceeds hu_max.

        Smoothing (closing) can push the segmentation boundary into cortical bone.
        This mask restores the HU constraint by zeroing out any voxel where the
        original CT intensity is above hu_max.
        """
        import vtk.util.numpy_support as nps # type: ignore
        import sitkUtils
        import SimpleITK as sitk

        seg = segmentation_node.GetSegmentation()
        binary_lm = seg.GetSegment(seg_id).GetRepresentation("Binary labelmap")
        if binary_lm is None:
            return

        # Export CT into the same geometry as the binary labelmap
        ref_lm = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLabelMapVolumeNode", "_TmpHURef")
        try:
            slicer.modules.segmentations.logic().ExportAllSegmentsToLabelmapNode(
                segmentation_node, ref_lm, slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY
            )
            ct_sitk  = sitkUtils.PullVolumeFromSlicer(volume_node)
            ref_sitk = sitkUtils.PullVolumeFromSlicer(ref_lm)
        finally:
            slicer.mrmlScene.RemoveNode(ref_lm)

        # Resample CT to labelmap geometry (nearest-neighbour, float keeps HU values)
        resampler = sitk.ResampleImageFilter()
        resampler.SetReferenceImage(ref_sitk)
        resampler.SetInterpolator(sitk.sitkLinear)
        resampler.SetDefaultPixelValue(-1000.0)
        ct_resampled = resampler.Execute(ct_sitk)

        ct_arr  = sitk.GetArrayFromImage(ct_resampled)   # (z, y, x)
        dims    = binary_lm.GetDimensions()               # (x, y, z)
        scalars = binary_lm.GetPointData().GetScalars()
        flat    = nps.vtk_to_numpy(scalars)
        seg_arr = flat.reshape(dims[2], dims[1], dims[0])  # (z, y, x)

        # Clip ct_arr to the labelmap extent if sizes differ by 1 (rounding)
        for axis, lm_size in enumerate(seg_arr.shape):
            if ct_arr.shape[axis] != lm_size:
                slices = [slice(None)] * 3
                slices[axis] = slice(0, lm_size)
                ct_arr = ct_arr[tuple(slices)]

        bone_mask = ct_arr > hu_max
        seg_arr[bone_mask] = 0

        scalars.DeepCopy(nps.numpy_to_vtk(seg_arr.astype(flat.dtype).flatten()))
        binary_lm.Modified()
        segmentation_node.Modified()

    def _removeSatelliteRegions(
        self, segmentation_node, seg_id, min_diameter_mm: float = 3.0
    ):
        """Removes connected components whose sphere-equivalent diameter is below min_diameter_mm.

        The largest component is always kept regardless of size.  For each other
        component, the effective diameter is computed as d = 2*(3V/4π)^(1/3), where V
        is the component volume in mm³.  Components below the threshold are deleted;
        the main body (largest component) is never touched.
        """
        import vtk.util.numpy_support as nps # pyright: ignore[reportMissingImports]
        from scipy import ndimage

        seg      = segmentation_node.GetSegmentation()
        binary_lm = seg.GetSegment(seg_id).GetRepresentation("Binary labelmap")
        if binary_lm is None:
            return

        dims    = binary_lm.GetDimensions()
        spacing = binary_lm.GetSpacing()
        scalars = binary_lm.GetPointData().GetScalars()

        flat      = nps.vtk_to_numpy(scalars)
        vol_array = flat.reshape(dims[2], dims[1], dims[0])
        binary    = vol_array > 0

        labeled, n_components = ndimage.label(binary)
        if n_components <= 1:
            print(f"    1 component – nothing to remove")
            return

        sizes       = np.array(ndimage.sum(binary, labeled, range(1, n_components + 1)))
        vox_vol_mm3 = spacing[0] * spacing[1] * spacing[2]
        largest     = int(np.argmax(sizes)) + 1

        keep_mask = np.zeros_like(binary)
        n_removed = 0
        for label_idx, size in enumerate(sizes, start=1):
            vol_mm3  = float(size) * vox_vol_mm3
            eff_diam = 2.0 * (3.0 * vol_mm3 / (4.0 * np.pi)) ** (1.0 / 3.0)
            if label_idx == largest or eff_diam >= min_diameter_mm:
                keep_mask |= (labeled == label_idx)
            else:
                n_removed += 1
                print(f"    Component {label_idx}: d_eff={eff_diam:.1f} mm → removed")

        print(f"    {n_components} components, {n_removed} removed")
        if n_removed > 0:
            flat_out = keep_mask.astype(np.uint8).flatten().astype(flat.dtype)
            scalars.DeepCopy(nps.numpy_to_vtk(flat_out))
            binary_lm.Modified()
            segmentation_node.Modified()

    def _prepareContralateralSeed(
        self, contra_seg_node, contra_seg_id, volume_node, name_prefix=""
    ):
        """Mirrors the contralateral segment, visualises both the raw mirror and a
        10 %-shrunk version as segmentation nodes, and returns the shrunk voxel
        indices as FM trial points.

        Returns a list of (i, j, k) index tuples (empty on failure).
        """
        import SimpleITK as sitk
        import sitkUtils

        # Export contralateral segment at full CT resolution
        seg_id_arr = vtk.vtkStringArray()
        seg_id_arr.InsertNextValue(contra_seg_id)
        tmp_lm = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode", "_TmpContraLM"
        )
        try:
            slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                contra_seg_node, seg_id_arr, tmp_lm, volume_node
            )
            contra_sitk = sitkUtils.PullVolumeFromSlicer(tmp_lm)
        finally:
            slicer.mrmlScene.RemoveNode(tmp_lm)

        contra_arr = sitk.GetArrayFromImage(contra_sitk)  # ZYX uint8

        # Mirror across L-R axis (axis=2 in ZYX = image i axis in standard LPS CT)
        mirrored = np.ascontiguousarray(np.flip(contra_arr, axis=2))
        print(f"  Contralateral mirror: {int(mirrored.sum()):,} voxels")

        # Display full mirrored mask (blue, semi-transparent)
        mirrored_sitk = sitk.GetImageFromArray(mirrored)
        mirrored_sitk.CopyInformation(contra_sitk)
        self._importBinaryAsSegNode(
            mirrored_sitk, volume_node,
            name=f"{name_prefix}ContraMirror_Full",
            color=(0.2, 0.6, 1.0),
            opacity=0.25,
        )

        # Collect nonzero indices as FM trial points
        nz_k, nz_j, nz_i = np.where(mirrored > 0)
        if len(nz_i) == 0:
            print("  Mirrored seed is empty – falling back to single seed.")
            return []

        return [(int(nz_i[m]), int(nz_j[m]), int(nz_k[m])) for m in range(len(nz_i))]

    def _prepareModelSeed(
        self, model_seg_node, model_seg_id, volume_node, name_prefix=""
    ):
        """Exports the user-positioned template segmentation into the CT volume's geometry
        and returns its voxel indices as FM trial points.

        The template is expected to have a vtkMRMLLinearTransformNode applied (via
        onPositionModelButton).  ExportSegmentsToLabelmapNode honours the parent
        transform automatically, so the exported labelmap reflects the positioned state.
        """
        import SimpleITK as sitk
        import sitkUtils

        # Gehärtete Kopie erstellen: ExportSegmentsToLabelmapNode berücksichtigt
        # Skalierungs-Transforms nicht immer korrekt; Härten bäckt die vollständige
        # Transform-Kette (Translation, Rotation, Skalierung) in die Geometrie ein.
        hardened_copy = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLSegmentationNode", "_TmpModelSegHardened"
        )
        hardened_copy.GetSegmentation().DeepCopy(model_seg_node.GetSegmentation())
        tf_id = model_seg_node.GetTransformNodeID()
        if tf_id:
            hardened_copy.SetAndObserveTransformNodeID(tf_id)
            slicer.modules.transforms.logic().hardenTransform(hardened_copy)

        seg_id_arr = vtk.vtkStringArray()
        seg_id_arr.InsertNextValue(model_seg_id)
        tmp_lm = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode", "_TmpModelLM"
        )
        try:
            slicer.modules.segmentations.logic().ExportSegmentsToLabelmapNode(
                hardened_copy, seg_id_arr, tmp_lm, volume_node
            )
            model_sitk = sitkUtils.PullVolumeFromSlicer(tmp_lm)
        finally:
            slicer.mrmlScene.RemoveNode(tmp_lm)
            slicer.mrmlScene.RemoveNode(hardened_copy)

        model_arr = sitk.GetArrayFromImage(model_sitk)
        n_vox = int(model_arr.sum())
        print(f"  Model seed: {n_vox:,} voxels")

        if n_vox == 0:
            print("  Model seed is empty – verify the template is positioned inside the CT volume.")
            return []

        # Display the full positioned template (purple) for reference
        self._importBinaryAsSegNode(
            model_sitk, volume_node,
            name=f"{name_prefix}ModelSeed",
            color=(0.7, 0.2, 0.9),
            opacity=0.25,
        )

        nz_k, nz_j, nz_i = np.where(model_arr > 0)
        return [(int(nz_i[m]), int(nz_j[m]), int(nz_k[m])) for m in range(len(nz_i))]

    def _importBinaryAsSegNode(self, binary_sitk, volume_node, name, color, opacity=0.35):
        """Creates (or replaces) a named segmentation node from a SimpleITK binary image."""
        import sitkUtils

        # Remove stale node from a previous run so the scene stays tidy
        old = slicer.mrmlScene.GetFirstNodeByName(name)
        if old is not None:
            slicer.mrmlScene.RemoveNode(old)

        tmp_lm = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode", "_TmpImportLM"
        )
        try:
            sitkUtils.PushVolumeToSlicer(binary_sitk, tmp_lm)
            seg_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode", name)
            seg_node.CreateDefaultDisplayNodes()
            seg_node.SetReferenceImageGeometryParameterFromVolumeNode(volume_node)
            slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
                tmp_lm, seg_node
            )
        finally:
            slicer.mrmlScene.RemoveNode(tmp_lm)

        seg = seg_node.GetSegmentation()
        if seg.GetNumberOfSegments() > 0:
            segment = seg.GetNthSegment(0)
            segment.SetColor(*color)
            segment.SetName(name)

        disp = seg_node.GetDisplayNode()
        disp.SetOpacity3D(opacity)
        disp.SetOpacity2DFill(opacity)
        disp.SetVisibility3D(True)
        seg_node.CreateClosedSurfaceRepresentation()

        return seg_node
    
    def _activeParameterNodeIndex(self, active_mrml_node=None) -> int:
        """Returns the 0-based index of the given PN among all module PNs, or -1."""
        if active_mrml_node is None:
            return -1
        all_parameter_nodes = self.getAllParameterNodes()
        for i, pn in enumerate(all_parameter_nodes):
            if active_mrml_node.GetID() == pn.GetID():
                return i
        return -1

    def _getColorForActiveParameterNode(self, objectType, active_mrml_node=None):
        idx = self._activeParameterNodeIndex(active_mrml_node)
        if idx == -1:
            return self.COLOR_DEFAULT

        colors = (self.COLORS_SEGMENTATION if objectType == "Segmentation"
                  else self.COLORS_ORBITAL_PLANE if objectType == "OrbitalPlane"
                  else [])
        if not colors or idx >= len(colors):
            return self.COLOR_DEFAULT
        return colors[idx]

    def resampleVolume(self, volume_node, target_spacing_mm: float):
        """Resample volume_node to isotropic target_spacing_mm using SimpleITK.

        Returns the new vtkMRMLScalarVolumeNode, or None on failure.
        """
        import SimpleITK as sitk
        from slicer.util import getNode
        import sitkUtils

        sitk_image = sitkUtils.PullVolumeFromSlicer(volume_node)
        original_spacing = sitk_image.GetSpacing()
        original_size = sitk_image.GetSize()

        new_spacing = (target_spacing_mm, target_spacing_mm, target_spacing_mm)
        new_size = [
            int(round(original_size[i] * original_spacing[i] / target_spacing_mm))
            for i in range(3)
        ]

        resampler = sitk.ResampleImageFilter()
        resampler.SetOutputSpacing(new_spacing)
        resampler.SetSize(new_size)
        resampler.SetOutputDirection(sitk_image.GetDirection())
        resampler.SetOutputOrigin(sitk_image.GetOrigin())
        resampler.SetTransform(sitk.Transform())
        resampler.SetDefaultPixelValue(-1024)  # air HU for out-of-bounds voxels
        resampler.SetInterpolator(sitk.sitkLinear)

        resampled_sitk = resampler.Execute(sitk_image)
        del sitk_image

        new_name = f"{volume_node.GetName()}_iso{target_spacing_mm:.2f}mm"
        new_node = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLScalarVolumeNode", new_name)
        sitkUtils.PushVolumeToSlicer(resampled_sitk, new_node)
        new_node.CreateDefaultDisplayNodes()
        return new_node